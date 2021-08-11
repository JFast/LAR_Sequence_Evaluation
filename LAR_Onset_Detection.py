import cv2
import numpy as np
import numpy.polynomial.polynomial as poly
import math
import time
from matplotlib import pyplot as plt
import DisplayVideo as display
import Preprocessing as prepro
import ReferencePoint as reference
import Segmentation as segmentation
import VocalFolds as vocalfold
import Fitting as fit
import Plots as plot
import UserInteraction as user
import openpyxl as pxl

# USER INTERACTION

# mouse callback
def callbackMouseClick(event, x, y, flags, param):
    if event == cv2.EVENT_LBUTTONDBLCLK:
        changeMask(x, y)


# trackbar interaction
def nothing(x):
    pass


# mask adaptation
def changeMask(x, y):
    # initialize point mask (point values at position 0, 5, 10, 15, ... set to 255, other values set to 0)
    mask = user.getMaskForUser(frame.shape[0], frame.shape[1])
    # "index" contains only point locations with value 255 (grid mask with steps of 5 pixels)
    index = np.where(mask == 255)
    # initialize "minimum" with length of frame diagonal
    minimum = math.sqrt(pow(abs(256), 2) + pow(abs(256), 2))
    # set "point" to "no point selected"
    point = None
    # for all grid mask points along first frame dimension
    for i in range(0, len(index[0])):
        # calculate distance of selected point in frame to current grid point
        distance = math.sqrt(pow(abs(index[0][i] - y), 2) + pow(abs(index[1][i] - x), 2))
        # if distance to current grid point lower than length of frame diagonal/"minimum" (selected point is valid)
        if distance < minimum:
            # update "minimum" to distance of selected point (x,y) to current grid point index[0][i], index[1][i]
            minimum = distance
            # set "point" to current grid point
            point = [index[0][i], index[1][i]]
    # show closest grid point
    print(point)
    # if closest point could be identified
    if not (point is None):
        # if value at this point is equal to 255
        if frame[point[0], point[1]][2] == 255:
            # set color of this point to green
            frame[point[0], point[1]] = [0, 255, 0]
            # set seed point
            mask_seeds[point[0], point[1]] = 255
        # if value at this point is not 255
        else:
            # set color of this point to red
            frame[point[0], point[1]] = [0, 0, 255]
            # do not set seed point
            mask_seeds[point[0], point[1]] = 0


# PATH DEFINITIONS
patient = "05"
sequence_number = "01"
# SPREADSHEET DEFINITIONS
spreadsheet_row = 52
# selection of glottal orientation correction method (PCA/iterative method)
# mode_orientation_correction = "PCA"
mode_orientation_correction = "iterative"
# use avi file
video_path = r"F:/LARvideos/videos_annotated/pat_" + patient + "/es_01_pat_" + patient + "_seq_" + sequence_number + \
             "/es_01_pat_" + patient + "_seq_" + sequence_number + ".avi"
# use mp4 file
# video_path = r"F/LARvideos/videos_annotated/pat_" + patient + "/es_01_pat_" + patient + "_seq_" + sequence_number +
# "/es_01_pat_" + patient + "_seq_" + sequence_number + ".mp4"

# output file path
saving_path = r"F:/Masterarbeit_Andra_Oltmann/Results_TMI/LAR_Onset_Detection/"

# spreadsheet file path
spreadsheet_path = 'F:/Masterarbeit_Andra_Oltmann/Results_TMI/Overview_Evaluation_Results.xlsx'

# open spreadsheet
spreadsheet = pxl.load_workbook(spreadsheet_path)
sheet = spreadsheet["Overview"]

file = open(saving_path + patient + "_" + sequence_number + "_Evaluation.txt", "w")
file.write("GENERAL INFORMATION\n")
file.write("Sequence identifier: es_01_pat_" + patient + "_seq_" + sequence_number + "\n")

# IDENTIFICATION OF INTRAGLOTTAL REFERENCE POINT

# INTENSITY VARIATION OVER FRAME COLUMNS
# get start time
start = time.time()
# load video file
video = display.loadVideo(video_path)
# variable for current frame index
frame_number = 0
# declare matrix S of column intensity variation (size not specified)
intensity_matrix_columns = np.zeros((0, 0))

while video.isOpened():
    # load next frame
    ret, frame = video.read()
    if not ret:
        # stop loop if no frame could be loaded (end of video)
        break
    else:
        # determine current frame number
        frame_number = frame_number + 1
        print(frame_number)

        # conversion into V channel of HSV color space and cropping to size of (216x216) px
        frame_v = prepro.convertToVChannel(frame[20:frame.shape[0]-20, 20:frame.shape[1]-20, :])
        # contrast enhancement of V channel
        frame_v = prepro.enhancementContrast(frame_v, 1.5)
        # transform image (matrix) into array
        frame_v_array = prepro.convertImageToArray(frame_v)

        # determine average intensity per column in frame (yields row vector)
        intensity_columns = reference.calculateIntensityColumn(frame_v_array)

        # add array with average intensity per column to intensity variation matrix (appends new row to matrix)
        intensity_matrix_columns = reference.getIntensityMatrix(intensity_matrix_columns, intensity_columns)

        # show V channel of original frame to inform user about progress
        frame_v_large = cv2.resize(frame_v, (int(2.0 * frame_v.shape[1]), int(2.0 * frame_v.shape[0])),
                                   interpolation=cv2.INTER_LINEAR)
        display.displayFrame("Current Frame", frame_v_large, 1)

# calculate average intensity differences along frame columns over all frames (yields row vector)
average_intensity = reference.calculateAverageIntensity(intensity_matrix_columns)
# calculate total intensity variations along frame columns over all frames (yields row vector)
total_intensity = reference.calculateTotalIntensity(intensity_matrix_columns, average_intensity)

# create array with column indices (maximum: number of frames in sequence)
columns = range(0, total_intensity.shape[1])
# fit Gaussian distribution to total intensity variation
# if adaptation not possible: do not crop frame in next step (set sigma to 256/2 = 128)
try:
    H, A, x0, sigma = reference.fitGaussianFunction(columns, total_intensity[0, :], 30)
except:
    sigma = 128
    x0 = np.argmax(total_intensity[0, :])

# evaluate fit values
sigma = reference.checkSigma(sigma, 128)
x0 = reference.checkX0(x0, total_intensity)
print(sigma, x0)
plot.plotTotalIntensity(total_intensity, "column index", saving_path + patient + "_" + sequence_number +
                          "_Intensity_Variation_Columns.png")
file.write("Total column intensity variation: \n")
file.write(str(total_intensity))
file.write("\n")

# for later use
# sigma of rows
sigma_x = sigma

# shift x0 for compliance with original frame size
x = int(x0) + 20

# initialize frame cropping for total intensity variation calculation over frame rows
min_boundary_columns, max_boundary_columns = reference.getBoundaries(x0, sigma, 2, 2)

# shift boundaries for compliance with original frame size
min_boundary_columns = min_boundary_columns + 20
max_boundary_columns = max_boundary_columns + 20
min_boundary_columns, max_boundary_columns = reference.checkBoundaries(min_boundary_columns, max_boundary_columns)

# close all windows
display.destroyWindows()

# INTENSITY VARIATION OVER FRAME ROWS

# load video file
video = display.loadVideo(video_path)
# variable for current frame index
frame_number = 0

# declare matrix S of frame row intensity variation (size not specified)
intensity_matrix_rows = np.zeros((0, 0))

while video.isOpened():
    # load next frame
    ret, frame = video.read()
    if not ret:
        # stop loop if no frame could be loaded (end of video)
        break
    else:
        # determine current frame index
        frame_number = frame_number + 1
        print(frame_number)

        # crop frame with boundary values identified during analysis of frame columns
        frame = frame[20:frame.shape[0] - 20, min_boundary_columns:max_boundary_columns]

        # conversion into HSV color space (V channel)
        frame_v = prepro.convertToVChannel(frame)
        # contrast enhancement of V channel
        frame_v = prepro.enhancementContrast(frame_v, 1.5)
        # transform frame (matrix) into array
        frame_v_array = prepro.convertImageToArray(frame_v)

        # average intensity per row
        intensity_rows = reference.calculateIntensityRow(frame_v_array)
        # add array with average intensity per row to intensity variation matrix
        intensity_matrix_rows = reference.getIntensityMatrix(intensity_matrix_rows, intensity_rows)

        # show V channel of original frame to inform user about progress
        frame_v_large = cv2.resize(frame_v, (int(2.0 * frame_v.shape[1]), int(2.0 * frame_v.shape[0])),
                                   interpolation=cv2.INTER_LINEAR)
        display.displayFrame("Current Frame", frame_v_large, 1)

# calculate average intensity differences along frame rows over all frames (yields row vector)
average_intensity = reference.calculateAverageIntensity(intensity_matrix_rows)
# calculate total intensity variations along frame rows over all frames (yields row vector)
total_intensity = reference.calculateTotalIntensity(intensity_matrix_rows, average_intensity)

# array with row indices
rows = range(0, total_intensity.shape[1])
# fit Gaussian distribution to total intensity variation
# if adaptation not possible: do not crop frame in next step (set sigma to 256/2 = 128)
try:
    H, A, x0, sigma = reference.fitGaussianFunction(rows, total_intensity[0, :], 30)
except:
    sigma = 128
    x0 = np.argmax(total_intensity[0, :])

print(sigma, x0)
plot.plotTotalIntensity(total_intensity, "row index", saving_path + patient + "_" + sequence_number +
                          "_Intensity_Variation_Rows.png")
file.write("Total row intensity variation: \n")
file.write(str(total_intensity))
file.write("\n")

# evaluate fit values
sigma = reference.checkSigma(sigma, 128)
x0 = reference.checkX0(x0, total_intensity)

# for later use
# sigma of columns
sigma_x = sigma

y = int(x0) + 20

# close all windows
display.destroyWindows()

# visualize result of ROI detection step
# load video file
video = display.loadVideo(video_path)
ret, frame = video.read()
# show reference point identification result
display.displayReferencePoint(frame, x, y, [0, 0, 255], saving_path + patient + "_" + sequence_number +
                              "_Reference_Point_First_Frame.png")

# close all windows
display.destroyWindows()

# store result
file.write("Reference Point: [" + str(x) + "," + str(y) + "]\n")


# SEMI-AUTOMATIC GLOTTIS SEGMENTATION IN FIRST FRAME OF SEQUENCE
# load video file
video = display.loadVideo(video_path)

# read first frame from sequence
ret, frame = video.read()

# initialize variable for frame index display
frame_number = 1

# initialize list of evaluated frames (image data)
frame_list = []

# store first frame in list
frame_list.append(frame)

# pre-processing

# bilateral filter (edge-preserving)
frame_bilateral = prepro.bilateralFiltering(frame, 5, 75, 75)

# transformation of RGB frame into grayscale frame
frame_gray = prepro.convertToGray(frame_bilateral)

# contrast enhancement
frame_gray = prepro.enhancementContrast(frame_gray, 1.1)
# frame_gray = prepro.applyCLAHE(frame_gray, 1.0)
# frame_gray = prepro.equalizeHistogram(frame_gray)

# bilateral filter (edge-preserving)
frame_gray = prepro.bilateralFiltering(frame_gray, 3, 25, 5)

# watershed segmentation

# automatic identification of Canny hysteresis threshold values
low_canny_thresh, high_canny_thresh = segmentation.getCannyThreshold(frame_gray, x, y)

# apply watershed segmentation with identified hysteresis threshold values
label = segmentation.getLabels(frame_gray, low_canny_thresh, high_canny_thresh)
watershed = segmentation.watershed_segmentation(cv2.cvtColor(frame_gray, cv2.COLOR_GRAY2BGR), label)
index_reference_point = segmentation.segment_glottis_point(watershed, x, y, frame.shape[0], frame.shape[1])
# display.displayWatershedSegmentation(frame, watershed, [0, 0, 255])

# get glottis contour from watershed result
glottis_contour_watershed = segmentation.getGlottisContourFromWatershed(frame.shape[0], frame.shape[1], watershed,
                                                                        index_reference_point)

# apply region growing (automated)
# get extremal points on glottis contour (watershed result)
extLeft, extRight, extTop, extBot = segmentation.getExtremePointsContour(glottis_contour_watershed)
# get point grid for region growing procedure (rectangular)
mask_grid_region_growing = segmentation.getGridForRegionGrowing(frame.shape[0], frame.shape[1],
                                                                extLeft, extRight, extTop, extBot)
# get seed point grid for region growing procedure (adapted to glottis contour)
seed_points = segmentation.getSeedPointsRegionGrowingFirstFrame(frame.shape[0], frame.shape[1],
                                                                glottis_contour_watershed, mask_grid_region_growing)
# calculate homogeneity criterion for region growing procedure
homogeneity_criterion = segmentation.getHomogeneityCriterion(frame_gray, seed_points)
# homogeneity_criterion = segmentation.getHomogeneityCriteriaRefined(frame_gray, seed_points)

# apply region growing procedure
region_growing = segmentation.regionGrowing(frame_gray, seed_points, homogeneity_criterion)
# region_growing = segmentation.regionGrowingRefined(frame_gray, seed_points, homogeneity_criterion)

# obtain glottis contour from region growing result
glottis_contour_region_growing = segmentation.getGlottisContourRegionGrowing(region_growing)

# let user check segmentation result in first frame of sequence
input_user_check = True
while input_user_check:
    if len(glottis_contour_region_growing) == 0:
        input_user = "n"
        print("Add seed points manually!")
    else:
        frame_contour = display.drawGlottisContour(frame, glottis_contour_region_growing, [0, 0, 255])
        input_user = input("Is the glottis segmentation correct? (y/n)\n")
    if input_user == "n":
        frame_result = frame.copy()
        mask_user = user.getMaskForUser(frame.shape[0], frame.shape[1])
        mask_seeds = np.zeros((frame.shape[0], frame.shape[1])).astype('uint8')
        frame[mask_user == 255] = [0, 0, 255]
        cv2.namedWindow('Input Mask', 1)
        cv2.setMouseCallback("Input Mask", callbackMouseClick)
        callback = True
        while callback:
            cv2.imshow('Input Mask', frame)
            k = cv2.waitKey(1) & 0xFF
            if k == ord('y'):
                seed_points = user.getSeedsFromMask(mask_seeds)
                # homogeneity_criterion = segmentation.getHomogeneityCriterion(frame_gray, seed_points)
                # homogeneity_criteria = segmentation.getHomogeneityCriteriaRefined(frame_gray, seed_points)
                # fix homogeneity criterion to value
                homogeneity_criterion = 1.02
                print("Homogeneity criterion: ", homogeneity_criterion)
                region_growing = segmentation.regionGrowing(frame_gray, seed_points, homogeneity_criterion)
                # region_growing = segmentation.regionGrowingRefined(frame_gray, seed_points, homogeneity_criteria)
                glottis_contour = segmentation.getGlottisContourRegionGrowing(region_growing)
                # calculate dense glottis contour for rotation identification
                glottis_contour_dense = segmentation.getGlottisContourRegionGrowingDense(region_growing)
                # add glottis contours to 'frame'
                frame = cv2.drawContours(frame, [glottis_contour], 0, [155, 88, 0], 1)
                cv2.imwrite('CorrectContour.png', frame)
                # draw glottis contours on 'frame_result' and display result
                display.drawGlottisContour(frame_result, glottis_contour, [0, 0, 255])
                callback = False
                file.write("Segmentation correct: no\n")
                file.write("Homogeneity criterion: ")
                file.write(str(homogeneity_criterion))
                file.write("\n")
        input_user_check = False
    elif input_user == "y":
        frame_result = frame.copy()
        input_user_check = False
        glottis_contour = glottis_contour_region_growing
        # calculate dense glottis contour for rotation identification via PCA
        glottis_contour_dense = segmentation.getGlottisContourRegionGrowingDense(region_growing)
        print("Identification of LAR onset time starting!")
        file.write("Segmentation correct: yes\n")
        file.write("Homogeneity criterion: ")
        file.write(str(homogeneity_criterion))
        file.write("\n")
    else:
        print("Input faulty!")

# IDENTIFICATION OF GLOTTAL PARAMETERS IN FIRST FRAME OF SEQUENCE

# calculate coordinates of frame center
x_center = 0.5 * frame.shape[1]
y_center = 0.5 * frame.shape[0]

# calculate centroid coordinates of glottis contour in first frame
M = cv2.moments(glottis_contour)
# horizontal coordinate of glottis centroid
cx = int(M['m10'] / M['m00'])
# vertical coordinate of glottis centroid
cy = int(M['m01'] / M['m00'])
print("cx of glottis centroid (non-dense contour): ", cx)
print("cy of glottis centroid (non-dense contour): ", cy)

# calculate glottal area
# 'glottal_area' is the glottal area in percent with respect to the total frame size in pixels
glottal_area = (cv2.contourArea(glottis_contour) / (frame.shape[0] * frame.shape[1])) * 100

# instantiate VideoWriter object
output_rotation = cv2.VideoWriter(
    saving_path + patient + "_" + sequence_number + '_rotation_correction_result.mp4',
    cv2.VideoWriter_fourcc(*"mp4v"), 15.0, (256, 256))

# instantiate VideoWriter object
output_rot_seg = cv2.VideoWriter(saving_path + patient + "_" + sequence_number +
                                 '_rotation_and_segmentation_result.mp4', cv2.VideoWriter_fourcc(*"mp4v"),
                                 15.0, (256, 256))

# instantiate VideoWriter object
output_all = cv2.VideoWriter(saving_path + patient + "_" + sequence_number +
                                 '_full_result.mp4', cv2.VideoWriter_fourcc(*"mp4v"),
                                 15.0, (256, 256))

# perform principal component analysis for identification of glottal midline orientation
if mode_orientation_correction == "PCA":
    # perform singular value decomposition to identify orientation of glottal midline in degrees
    # (required to rotate all frames for correct identification of glottal angle and vocal fold edge distance)
    angle_glottal_midline = segmentation.getGlottalOrientation(glottis_contour_dense)
    print("angle_glottal_midline in degrees: ", angle_glottal_midline)
# perform iterative procedure using relative horizontal position of vertex point
# for identification of glottal midline orientation
elif mode_orientation_correction == "iterative":
    # initialize variable for glottal midline inclination angle in degrees
    angle_glottal_midline_iterative = 0.0
    # initialize angle step for iterative angle identification in degrees
    delta_angle = 0.5
    # initialize initial and rotated glottis contours for iterative angle identification
    glottis_contour_initial = glottis_contour.copy()
    glottis_contour_iterative = glottis_contour.copy()
    # initialize rotated frame for iterative angle identification
    frame_initial = frame.copy()
    # initialize variables for rotation angle storage
    angle_before_cw = 0.0
    angle_before_ccw = 0.0
    while 1:
        # localization of points on vocal fold edges for calculation of glottal angle
        left_point_glottis, right_point_glottis, left_point_bottom, right_point_bottom = \
            vocalfold.getGlottalPoints(frame, glottis_contour_iterative)
        # identification of vertex point of glottal angle
        vertex_point = vocalfold.getVertexPoint(left_point_glottis, right_point_glottis, left_point_bottom,
                                                right_point_bottom)
        # calculation of horizontal position of vertex point in relation to left and right point for glottal angle
        # value < 0: vertex point left of left point; 0 < value < 1: vertex point between left and right point;
        # value > 1: vertex point right of right point
        relative_location_vertex_point_horizontal_iterative = (vertex_point[0] - left_point_glottis[0]) / \
                                                              (right_point_glottis[0] - left_point_glottis[0])
        print("relative_location_vertex_point_horizontal_iterative: ",
              relative_location_vertex_point_horizontal_iterative)
        # create window for display of rotated frame and glottis contour
        cv2.namedWindow("Rotated frame and glottis contour", cv2.WINDOW_AUTOSIZE)
        # if vertex located closer to left point on vocal fold edge
        if relative_location_vertex_point_horizontal_iterative < 0.48:
            # decrement glottal midline inclination angle in degrees
            angle_glottal_midline_iterative -= delta_angle
            # rotate glottis contour CCW by total rotation angle to correct rotation of glottal midline
            # for all points of initial glottis contour (no. of points in contour remains constant)
            for i in range(0, len(glottis_contour_initial)):
                # access current point
                point = glottis_contour_initial[i][0].copy()
                # rotate point around frame center
                point = segmentation.rotate_point((x_center, y_center), point, -angle_glottal_midline_iterative)
                # save rotated point
                glottis_contour_iterative[i][0][0] = point[0].copy()
                glottis_contour_iterative[i][0][1] = point[1].copy()
            # rotate 'frame_iterative'
            frame_iterative = segmentation.rotate_frame(frame_initial, -angle_glottal_midline_iterative)
            # draw rotated glottis contour on rotated frame
            frame_iterative = cv2.drawContours(frame_iterative, [glottis_contour_iterative], 0, [155, 88, 0], 1)
            # show frame
            cv2.imshow("Rotated frame and glottis contour", frame_iterative)
            # write frame to output sequences
            output_rotation.write(frame_iterative)
            output_rot_seg.write(frame_iterative)
            output_all.write(frame_iterative)

            key = cv2.waitKey(40) & 0xFF
            # if escape key was pressed
            if key == 27:
                cv2.destroyWindow("Rotated frame and glottis contour")
                angle_glottal_midline = 0.0
                break
            if angle_glottal_midline_iterative == angle_before_ccw:
                cv2.destroyWindow("Rotated frame and glottis contour")
                angle_glottal_midline = angle_glottal_midline_iterative + (0.5 * delta_angle)
                break
            angle_before_ccw = angle_glottal_midline_iterative
        # if vertex located closer to right point on vocal fold edge
        elif relative_location_vertex_point_horizontal_iterative > 0.52:
            # increment glottal midline inclination angle in degrees
            angle_glottal_midline_iterative += delta_angle
            # rotate glottis contour CW by total rotation angle to correct rotation of glottal midline
            # for all points of initial glottis contour (no. of points in contour remains constant)
            for i in range(0, len(glottis_contour_initial)):
                # access current point
                point = glottis_contour_initial[i][0].copy()
                # rotate point around frame center
                point = segmentation.rotate_point((x_center, y_center), point, -angle_glottal_midline_iterative)
                # save rotated point
                glottis_contour_iterative[i][0][0] = point[0].copy()
                glottis_contour_iterative[i][0][1] = point[1].copy()
            # rotate 'frame_iterative'
            frame_iterative = segmentation.rotate_frame(frame_initial, -angle_glottal_midline_iterative)
            # draw rotated glottis contour on rotated frame
            frame_iterative = cv2.drawContours(frame_iterative, [glottis_contour_iterative], 0, [155, 88, 0], 1)
            # show frame
            cv2.imshow("Rotated frame and glottis contour", frame_iterative)
            output_rotation.write(frame_iterative)
            output_rot_seg.write(frame_iterative)
            output_all.write(frame_iterative)
            key = cv2.waitKey(40) & 0xFF
            # if escape key was pressed
            if key == 27:
                cv2.destroyWindow("Rotated frame and glottis contour")
                angle_glottal_midline = 0.0
                break
            if angle_glottal_midline_iterative == angle_before_cw:
                cv2.destroyWindow("Rotated frame and glottis contour")
                angle_glottal_midline = angle_glottal_midline_iterative - (0.5 * delta_angle)
                break
            angle_before_cw = angle_glottal_midline_iterative
        # if vertex located centrally between left and right point of glottal angle (in horizontal direction)
        # ideal rotation correction angle found
        else:
            cv2.destroyWindow("Rotated frame and glottis contour")
            angle_glottal_midline = angle_glottal_midline_iterative
            print("angle_glottal_midline in degrees: ", angle_glottal_midline)
            break

# rotate frame to correct identified rotation of glottal midline
frame_rotation_check = segmentation.rotate_frame(frame, -angle_glottal_midline)

# show rotated frame and ask for user confirmation
cv2.imshow("First Frame (Rotated)", frame_rotation_check)
print("Rotation correction effective/glottal midline oriented vertically? Press 'n' to proceed without rotation or "
      "'m' for manual rotation angle adjustment. Any other input: accept identified rotation angle.")
k = cv2.waitKey(0) & 0xFF
if k == ord('n'):
    # update Boolean variable for rotation correction
    rotationCorrection = False
    # close window
    cv2.destroyWindow("First Frame (Rotated)")
    # write rotation correction angle to file
    file.write("Rotation correction by angle (in degrees, positive CCW): ")
    file.write(str(0))
    file.write("\n")
elif k == ord('m'):
    # close window
    cv2.destroyWindow("First Frame (Rotated)")
    # manual rotation angle adjustment
    print("Manual rotation adjustment started. Press return key to accept selected rotation angle. Displayed angle "
          "offset by +45 degrees.")
    # create window for rotated frame
    cv2.namedWindow("Rotated Frame")
    # initialize rotation angle (in degrees)
    angle_interactive = 0
    # create trackbar
    cv2.createTrackbar("ang+45deg", "Rotated Frame", 45, 90, nothing)
    while 1:
        # read trackbar value in degrees and subtract offset of +45 degrees
        angle_interactive = cv2.getTrackbarPos("ang+45deg", "Rotated Frame") - 45
        # show rotated frame
        frame_rotation_interactive = segmentation.rotate_frame(frame, angle_interactive)
        cv2.imshow("Rotated Frame", frame_rotation_interactive)
        # poll user input
        key = cv2.waitKey(1) & 0xFF
        # if return key pressed: save current rotation correction angle
        if key == 13:
            angle_glottal_midline = -angle_interactive
            break
    # write rotation correction angle to file
    file.write("Rotation correction by angle (in degrees, positive CCW): ")
    file.write(str(-angle_glottal_midline))
    file.write("\n")
    # close windows
    cv2.destroyWindow("Rotated Frame")
    # update Boolean variable for rotation correction
    rotationCorrection = True
else:
    # update Boolean variable for rotation correction
    rotationCorrection = True
    # close window
    cv2.destroyWindow("First Frame (Rotated)")
    # write rotation correction angle to file
    file.write("Rotation correction by angle (in degrees, positive CCW): ")
    file.write(str(-angle_glottal_midline))
    file.write("\n")

if rotationCorrection:
    # rotate frame to correct glottal orientation
    frame = segmentation.rotate_frame(frame, -angle_glottal_midline)
    frame_result = segmentation.rotate_frame(frame_result, -angle_glottal_midline)

    # rotate glottis contour around frame center
    # for all points in list 'glottis_contour'
    for point_ in glottis_contour:
        point = point_[0]
        # rotate point around frame center
        point = segmentation.rotate_point((x_center, y_center), point, -angle_glottal_midline)
    # rotate seed points accordingly
    for seed in seed_points:
        # rotate seed around frame center
        seed = segmentation.rotate_point((x_center, y_center), seed, -angle_glottal_midline)
    # rotate glottal reference point accordingly
    ref_rot = segmentation.rotate_point((x_center, y_center), [x, y], -angle_glottal_midline)

# localization of glottal points of interest for calculation of glottal angle
left_point_glottis, right_point_glottis, left_point_bottom, right_point_bottom = \
    vocalfold.getGlottalPoints(frame_result, glottis_contour)

# identify points for calculation of medial vocal fold distance
left_distance_point, right_distance_point = vocalfold.getDistancePoints(frame_result, glottis_contour, 0.5)
# save vertical position of medial vocal fold distance
constant_height_distance = left_distance_point[1]

print(left_point_glottis, right_point_glottis)

# calculate medial vocal fold distance in pixel
distance_between_glottis_points = abs(left_distance_point[0] - right_distance_point[0])

# identify vertex point for calculation of glottal angle
vertex_point = vocalfold.getVertexPoint(left_point_glottis, right_point_glottis, left_point_bottom, right_point_bottom)

# calculate glottal angle
glottal_angle = vocalfold.angle_straight_lines(left_point_glottis, right_point_glottis, vertex_point)

# calculation of horizontal position of vertex point in relation to left and right point for glottal angle
# (this relative location is set to remain CONSTANT throughout the frame sequence)
# value < 0: vertex point left of left point; 0 < value < 1: vertex point between left and right point;
# value > 1: vertex point right of right point
relative_location_vertex_point_horizontal = (vertex_point[0] - left_point_glottis[0]) \
                                            / (right_point_glottis[0] - left_point_glottis[0])

# frame_result = frame.copy()

# draw points for glottal angle
frame_result = cv2.circle(frame_result, (int(left_point_glottis[0]), int(left_point_glottis[1])), 2, [255, 0, 0], -1)
frame_result = cv2.circle(frame_result, (int(right_point_glottis[0]), int(right_point_glottis[1])), 2, [255, 0, 0], -1)
frame_result = cv2.circle(frame_result, (int(vertex_point[0]), int(vertex_point[1])), 2, [255, 0, 0], -1)
# draw points for vocal fold edge distance
frame_result = cv2.circle(frame_result, (int(left_distance_point[0]), int(left_distance_point[1])), 2, [255, 255, 0], -1)
frame_result = cv2.circle(frame_result, (int(right_distance_point[0]), int(right_distance_point[1])),
                          2, [255, 255, 0], -1)
# draw lines for glottal angle
frame_result = cv2.line(frame_result, (int(left_point_glottis[0]), int(left_point_glottis[1])),
                        (int(vertex_point[0]), int(vertex_point[1])), [255, 0, 0], 1)
frame_result = cv2.line(frame_result, (int(right_point_glottis[0]), int(right_point_glottis[1])),
                        (int(vertex_point[0]), int(vertex_point[1])), [255, 0, 0], 1)
# draw glottal angle
axis_angle = vocalfold.getAngleBetweenPoints([0, vertex_point[1]], vertex_point, left_point_glottis)
frame_result = cv2.ellipse(frame_result,
                           (int(vertex_point[0]), int(vertex_point[1])),
                           (int((vertex_point[1] - left_point_glottis[1]) / 2.0),
                            int((vertex_point[1] - left_point_glottis[1]) / 2.0)), 180.0 + axis_angle, 0.0,
                           glottal_angle, [255, 0, 0], 1)

# draw line for vocal fold edge distance
frame_result = cv2.line(frame_result, (int(left_distance_point[0]), int(left_distance_point[1])),
                        (int(right_distance_point[0]), int(right_distance_point[1])), [255, 255, 0], 1)
# draw glottis segmentation result
frame_result = cv2.drawContours(frame_result, [glottis_contour], 0, [0, 255, 0], 1)

frame_result_large = cv2.resize(frame_result, (int(2.0 * frame_result.shape[1]), int(2.0 * frame_result.shape[0])),
                                interpolation=cv2.INTER_LINEAR)
display.displayFrame("Segmentation Result (First Frame)", frame_result_large, 0)
cv2.imwrite(saving_path + patient + "_" + sequence_number + "_Glottal_Segmentation_First_Frame.png", frame_result)
for i in range(0, 30):
    output_rotation.write(frame_result)
    output_rot_seg.write(frame_result)
    output_all.write(frame_result)
    i += 1
output_rotation.release()
# add metrics to lists
frame_number_list_distance = list()
distance_list = list()
frame_number_list_angle = list()
angle_list = list()
frame_number_list_area = list()
area_list = list()

# add frame number '1' to lists
frame_number_list_distance.append(frame_number)
frame_number_list_angle.append(frame_number)
frame_number_list_area.append(frame_number)

distance_list.append(distance_between_glottis_points)

# "angle_list" contains temporal evolution of glottal angle
angle_list.append(glottal_angle)

# "area_list" contains temporal evolution of glottal area in percent w. r. t. total frame size in pixels
area_list.append(glottal_area)
area = cv2.contourArea(glottis_contour)

# initialize counter for number of unsuccessful glottis segmentation attempts
check_mean = 0

mid_point = left_distance_point[0] + (abs(left_distance_point[0] - right_distance_point[0])/2.0)

# initialize glottis contour in last iteration
glottis_contour_before = glottis_contour
# initialize glottis contour in second-last iteration
glottis_contour_before_before = glottis_contour

# SEGMENTATION OF FURTHER FRAMES (SECOND FRAME AND ON)

# instantiate VideoWriter object
output = cv2.VideoWriter(saving_path + patient + "_" + sequence_number + '_segmentation_result.mp4',
                         cv2.VideoWriter_fourcc(*"mp4v"), 15.0, (256, 256))

while video.isOpened():
    # load next frame
    ret, frame = video.read()
    if not ret:
        # stop loop if no frame could be loaded (end of video)
        break
    else:
        frame_list.append(frame)
        # start with frame_number = 2
        frame_number = frame_number + 1
        print(frame_number)

        if rotationCorrection:
            # rotate frame to correct glottal orientation
            frame = segmentation.rotate_frame(frame, -angle_glottal_midline)

        frame_result = frame.copy()

        # pre-processing

        # bilateral filter (edge-preserving)
        frame_bilateral = prepro.bilateralFiltering(frame, 5, 75, 75)
        # transformation of RGB frame into grayscale frame
        frame_gray = prepro.convertToGray(frame_bilateral)
        # contrast enhancement
        frame_gray = prepro.enhancementContrast(frame_gray, 1.1)
        # frame_gray = prepro.applyCLAHE(frame_gray, 1.0)
        # frame_gray = prepro.equalizeHistogram(frame_gray)
        # bilateral filter (edge-preserving)
        frame_gray = prepro.bilateralFiltering(frame_gray, 3, 25, 5)

        # region growing

        # analyze every 4th frame only --> virtual frame rate of 1000 fps
        if divmod(frame_number-1, 4)[1] == 0:
            # if glottis contour exists
            if not (len(glottis_contour) == 0):
                # get extremal points on glottis contour (using segmentation result of previous frame)
                extLeft, extRight, extTop, extBot = segmentation.getExtremePointsContour(glottis_contour)
                # get rectangular point grid for region growing procedure
                mask_grid_region_growing = segmentation.getGridForRegionGrowing(frame.shape[0],
                                                                                frame.shape[1],
                                                                                extLeft,
                                                                                extRight,
                                                                                extTop,
                                                                                extBot)
                # get seed point grid for region growing procedure (adapted to glottis contour)
                seed_points = segmentation.getSeedPointsRegionGrowingFirstFrame(frame.shape[0],
                                                                                frame.shape[1],
                                                                                glottis_contour,
                                                                                mask_grid_region_growing)
                # if no seed points available
                if len(seed_points) == 0:
                    # use glottis centroid as single seed point
                    seed_points = [[cx, cy]]

            # if less than two unsuccessful attempts of region growing have been completed
            if check_mean < 2:
                # apply region growing procedure
                region_growing = segmentation.regionGrowing(frame_gray, seed_points, homogeneity_criterion)
            else:
                # set region growing result to empty set if more than one unsuccessful region growing attempt occurred
                # (this stops glottis segmentation procedure)
                region_growing = np.zeros((frame.shape[0], frame.shape[1])).astype('uint8')

            # identify glottis contour after region growing procedure
            glottis_contour = segmentation.getGlottisContourRegionGrowing(region_growing)

            # if no contour found (empty list returned)
            if glottis_contour == []:
                # increment counter of repetitions
                check_mean = check_mean + 1

            # if glottis contour available
            if not glottis_contour == []:
                # if glottal area not zero
                if not area == 0:
                    # if glottis area has increased by more than 50% with respect to previous frame
                    # (occurs when glottis closure completed)
                    if cv2.contourArea(glottis_contour)/area > 1.50:
                        print("area")
                        # increment 'check_mean'
                        check_mean = check_mean + 1
                        # set glottis contour to empty set
                        glottis_contour = []
                        # if exactly one unsuccessful segmentation attempt completed
                        if check_mean == 1:
                            # set glottis segmentation result of current frame n to result of frame (n-2)
                            glottis_contour = glottis_contour_before_before

            # IDENTIFICATION OF GLOTTAL PARAMETERS
            # if list 'glottis_contour' not empty
            if not len(glottis_contour) == 0:
                try:
                    print("distance")
                    # localization of points for vocal fold distance calculation (with dynamic height adaptation)
                    # left_distance_point, right_distance_point = vocalfold.getDistancePoints(frame,
                    #                                                                         glottis_contour, 0.5)
                    # localization of points for vocal fold distance calculation (at constant height in frame)
                    left_distance_point, right_distance_point = vocalfold.getDistancePointsConstantHeight(
                        frame, glottis_contour, constant_height_distance)
                    # calculation of vocal fold distance
                    distance_between_glottis_points = abs(left_distance_point[0] - right_distance_point[0])

                    # calculation of horizontal coordinate of central point between points on medial vocal folds
                    # mid_point_new = left_distance_point[0] + (abs(left_distance_point[0] -
                    # right_distance_point[0]) / 2.0)
                    # if new central point has shifted
                    # if abs(mid_point_new - mid_point) >= 1:
                    #     # shift vertex point horizontally to compensate horizontal shift of glottis in image
                    #     vertex_point = [vertex_point[0] + (mid_point_new - mid_point), vertex_point[1]]
                    #     mid_point = mid_point_new

                    # add new distance value to list 'distance_list'
                    frame_number_list_distance.append(frame_number)
                    distance_list.append(distance_between_glottis_points)
                    # draw points on medial vocal folds
                    frame = cv2.circle(frame, (int(left_distance_point[0]), int(left_distance_point[1])),
                                       2, [255, 255, 0], -1)
                    frame = cv2.circle(frame, (int(right_distance_point[0]), int(right_distance_point[1])),
                                       2, [255, 255, 0], -1)
                    # draw distance line between vocal folds
                    frame = cv2.line(frame, (int(left_distance_point[0]), int(left_distance_point[1])),
                                     (int(right_distance_point[0]), int(right_distance_point[1])), [255, 255, 0], 1)
                except:
                    print("pass distance")

                try:
                    # points for glottal angle calculation do not move vertically
                    left_point_glottis = vocalfold.getPointOnGlottisContour(frame.shape[1], frame.shape[0],
                                                                            glottis_contour, left_point_glottis)
                    right_point_glottis = vocalfold.getPointOnGlottisContour(frame.shape[1], frame.shape[0],
                                                                             glottis_contour, right_point_glottis)
                    print(left_point_glottis, right_point_glottis)

                    if left_point_glottis == []:
                        pass
                    elif right_point_glottis == []:
                        pass
                    else:
                        print("angle")

                        # update horizontal position of vertex point with respect to left and right point
                        if right_point_glottis[0] > left_point_glottis[0]:
                            vertex_point = [int(round(left_point_glottis[0] +
                                                      relative_location_vertex_point_horizontal *
                                                      (right_point_glottis[0] - left_point_glottis[0]))),
                                            vertex_point[1]]

                        # calculate glottal angle using three defining points
                        glottal_angle = vocalfold.angle_straight_lines(left_point_glottis, right_point_glottis,
                                                                       vertex_point)
                        # add new angle value to list
                        frame_number_list_angle.append(frame_number)
                        angle_list.append(glottal_angle)
                        # draw points on vocal fold edges
                        frame = cv2.circle(frame, (int(left_point_glottis[0]), int(left_point_glottis[1])),
                                           2, [255, 0, 0], -1)
                        frame = cv2.circle(frame, (int(right_point_glottis[0]), int(right_point_glottis[1])),
                                           2, [255, 0, 0], -1)
                        # draw vertex point
                        frame = cv2.circle(frame, (int(vertex_point[0]), int(vertex_point[1])), 2, [255, 0, 0], -1)

                        # draw lines defining glottal angle
                        frame = cv2.line(frame, (int(left_point_glottis[0]), int(left_point_glottis[1])),
                                         (int(vertex_point[0]), int(vertex_point[1])), [255, 0, 0], 1)
                        frame = cv2.line(frame, (int(right_point_glottis[0]), int(right_point_glottis[1])),
                                         (int(vertex_point[0]), int(vertex_point[1])), [255, 0, 0], 1)
                        # draw glottal angle
                        axis_angle = vocalfold.getAngleBetweenPoints([0, vertex_point[1]], vertex_point,
                                                                     left_point_glottis)
                        frame = cv2.ellipse(frame, (int(vertex_point[0]), int(vertex_point[1])),
                                            (int((vertex_point[1] - left_point_glottis[1]) / 2.0),
                                             int((vertex_point[1] - left_point_glottis[1]) / 2.0)),
                                            180.0 + axis_angle, 0.0, glottal_angle, [255, 0, 0], 1)
                except:
                    print("pass angle")

                try:
                    # calculate glottal area in percent of total frame area
                    glottal_area = (cv2.contourArea(glottis_contour) / (frame.shape[0] * frame.shape[1])) * 100
                    # add new area percentage value to list
                    frame_number_list_area.append(frame_number)
                    area_list.append(glottal_area)
                except:
                    pass

                # update glottis centroid coordinates
                M = cv2.moments(glottis_contour)
                if not (M['m00'] == 0):
                    # horizontal coordinate of glottis centroid
                    cx = int(M['m10']/M['m00'])
                    # vertical coordinate of glottis centroid
                    cy = int(M['m01']/M['m00'])

                # calculate glottal area
                area = cv2.contourArea(glottis_contour)
                # if fourth frame in analysis (frame_number equal to 13)
                # if second frame in analysis (frame_number equal to 5)
                if frame_number == 5:
                    # store current glottis segmentation result for comparison with future segmentations
                    glottis_contour_before = glottis_contour
                # if fifth frame in analysis or later (frame_number larger or equal 17)
                # if third frame in analysis or later (frame_number larger or equal 9)
                if frame_number >= 9:
                    # update glottis segmentation result from two frames before for comparison with future segmentations
                    glottis_contour_before_before = glottis_contour_before
                    # update glottis segmentation result from last frame for comparison with future segmentations
                    glottis_contour_before = glottis_contour
                # draw glottis contour
                frame = cv2.drawContours(frame, [glottis_contour], 0, [0, 255, 0], 1)
                frame_result = cv2.drawContours(frame_result, [glottis_contour], 0, [0, 255, 0], 1)
            frame_large = cv2.resize(frame, (int(2.0*frame.shape[1]), int(2.0*frame.shape[0])),
                                     interpolation=cv2.INTER_LINEAR)
            cv2.imshow("Current Segmentation Result", frame_large)
            cv2.waitKey(1)
            # save frame showing glottis segmentation result
            output.write(frame)
            output_rot_seg.write(frame)
            output_all.write(frame)
output.release()
output_rot_seg.release()
end = time.time()

# DATA PROCESSING

# write raw data points to evaluation document
file.write("\n")
file.write("DATA\n\n")
file.write("Evolution of vocal fold edge distance: \n")
file.write(str(frame_number_list_distance))
file.write("\n")
file.write(str(distance_list))
file.write("\n")
file.write("Evolution of glottal angle: \n")
file.write(str(frame_number_list_angle))
file.write("\n")
file.write(str(angle_list))
file.write("\n")
file.write("Evolution of glottal area: \n")
file.write(str(frame_number_list_area))
file.write("\n")
file.write(str(area_list))
file.write("\n\n")

# FITTING AND RESULT VISUALIZATION

file.write("EVALUATION\n\n")
file.write("Computation time in s: " + '{:5.3f} s'.format(end-start) + "\n\n")
file.write("LAR ONSET TIME IDENTIFICATION\n\n")

# VOCAL FOLD EDGE DISTANCE

file.write("VOCAL FOLD EDGE DISTANCE\n\n")

# symmetrical sigmoid fit without vertical offset (distance)
try:
    x_distance_sigmoid, y_distance_sigmoid, popt_distance_sigmoid = fit.sigmoid_fit(frame_number_list_distance,
                                                                                    distance_list)
    sigmoid_distance_95 = fit.getValueNoVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                       popt_distance_sigmoid, 0.95)
    sigmoid_distance_96 = fit.getValueNoVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                       popt_distance_sigmoid, 0.96)
    sigmoid_distance_97 = fit.getValueNoVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                       popt_distance_sigmoid, 0.97)
    sigmoid_distance_98 = fit.getValueNoVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                       popt_distance_sigmoid, 0.98)
    sigmoid_distance_99 = fit.getValueNoVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                       popt_distance_sigmoid, 0.99)

    plot.plotSigmoidFitDistance(frame_number_list_distance, distance_list, x_distance_sigmoid, y_distance_sigmoid,
                                saving_path + patient + "_" + sequence_number +
                                "_Symmetrical_Sigmoid_Vocal_Fold_Edge_Distance.png")

    print("Vocal fold edge distance (symm. sigmoid without vertical offset):",
          [sigmoid_distance_95, sigmoid_distance_96, sigmoid_distance_97, sigmoid_distance_98, sigmoid_distance_99])

    file.write("LAR onset frame via vocal fold edge distance (symm. sigmoid without vertical offset, "
               "95/96/97/98/99% decline): ")
    file.write(str([sigmoid_distance_95, sigmoid_distance_96, sigmoid_distance_97,
                    sigmoid_distance_98, sigmoid_distance_99]))
    file.write("\n")
    file.write("Parameters (symm. sigmoid without vertical offset, vocal fold edge distance): ")
    file.write(str(popt_distance_sigmoid))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    # index of last data point used for vocal fold distance fitting
    last_index_fit_distance = frame_number_list_distance.index(len(x_distance_sigmoid) - 1)
    for i in range(0, last_index_fit_distance):
        error_sum += pow(abs(distance_list[i] - fit.sigmoid(frame_number_list_distance[i], *popt_distance_sigmoid)), 2)
    error_sum /= last_index_fit_distance
    rmse = math.sqrt(error_sum)
    file.write("RMSE (symm. sigmoid without vertical offset, vocal fold edge distance) in pixel: ")
    file.write(str(rmse))
    file.write("\n")

    # # use function for RMSE calculation
    # rmse = fit.getRMSE('sigmoid', frame_number_list_distance, distance_list, last_index_fit_distance, popt_distance_sigmoid)
    # file.write("RMSE (symm. sigmoid without vertical offset, vocal fold edge distance) in pixel: ")
    # file.write(str(rmse))
    # file.write("\n")

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_distance):
        error_sum_MAE += abs(distance_list[i] - fit.sigmoid(frame_number_list_distance[i], *popt_distance_sigmoid))
    error_sum_MAE /= last_index_fit_distance
    file.write("MAE (symm. sigmoid without vertical offset, vocal fold edge distance) in pixel: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # # use function for MAE calculation
    # mae = fit.getMAE('sigmoid', frame_number_list_distance, distance_list, last_index_fit_distance, popt_distance_sigmoid)
    # file.write("MAE (symm. sigmoid without vertical offset, vocal fold edge distance) in pixel: ")
    # file.write(str(mae))
    # file.write("\n\n")
except:
    file.write("Vocal fold edge distance: symm. sigmoid fit without vertical offset not successful!\n\n")

# symmetrical sigmoid fit with vertical offset (distance)
try:
    x_distance_sigmoid, y_distance_sigmoid, popt_distance_sigmoid_offset = \
        fit.sigmoid_fit_offset(frame_number_list_distance, distance_list)
    sigmoid_distance_95 = fit.getValueWithVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                         popt_distance_sigmoid_offset, 0.95)
    sigmoid_distance_96 = fit.getValueWithVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                         popt_distance_sigmoid_offset, 0.96)
    sigmoid_distance_97 = fit.getValueWithVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                         popt_distance_sigmoid_offset, 0.97)
    sigmoid_distance_98 = fit.getValueWithVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                         popt_distance_sigmoid_offset, 0.98)
    sigmoid_distance_99 = fit.getValueWithVerticalOffset(x_distance_sigmoid, y_distance_sigmoid,
                                                         popt_distance_sigmoid_offset, 0.99)

    plot.plotSigmoidFitDistanceOffset(frame_number_list_distance, distance_list, x_distance_sigmoid, y_distance_sigmoid,
                                saving_path + patient + "_" + sequence_number +
                                "_Symmetrical_Sigmoid_Vertical_Offset_Vocal_Fold_Edge_Distance.png")

    print("Vocal fold edge distance (symm. sigmoid with vertical offset): ",
          [sigmoid_distance_95, sigmoid_distance_96, sigmoid_distance_97, sigmoid_distance_98, sigmoid_distance_99])

    file.write("LAR onset frame via vocal fold edge distance (symm. sigmoid fit with vertical offset, "
               "95/96/97/98/99% decline): ")
    file.write(str([sigmoid_distance_95, sigmoid_distance_96, sigmoid_distance_97, sigmoid_distance_98,
                    sigmoid_distance_99]))
    file.write("\n")
    file.write("Parameters (symm. sigmoid with vertical offset, vocal fold edge distance): ")
    file.write(str(popt_distance_sigmoid_offset))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_distance):
        error_sum += pow(abs(distance_list[i] - fit.sigmoid_offset(frame_number_list_distance[i],
                                                                   *popt_distance_sigmoid_offset)), 2)
    error_sum /= last_index_fit_distance
    rmse = math.sqrt(error_sum)
    file.write("RMSE (symm. sigmoid with vertical offset, vocal fold edge distance) in pixel: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=32).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_distance):
        error_sum_MAE += abs(distance_list[i] - fit.sigmoid_offset(frame_number_list_distance[i],
                                                                   *popt_distance_sigmoid_offset))
    error_sum_MAE /= last_index_fit_distance
    file.write("MAE (symm. sigmoid with vertical offset, vocal fold edge distance) in pixel: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=33).value = error_sum_MAE
except:
    file.write("Vocal fold edge distance: symm. sigmoid fit with vertical offset not successful!\n\n")

# generalized logistic function (distance)
try:
    x_distance_glf, y_distance_glf, popt_distance_glf = \
        fit.fit_generalized_logistic_function(frame_number_list_distance, distance_list)
    glf_distance_95 = fit.getValueWithVerticalOffset(x_distance_glf, y_distance_glf,
                                                 popt_distance_glf, 0.95)
    glf_distance_96 = fit.getValueWithVerticalOffset(x_distance_glf, y_distance_glf,
                                                 popt_distance_glf, 0.96)
    glf_distance_97 = fit.getValueWithVerticalOffset(x_distance_glf, y_distance_glf,
                                                 popt_distance_glf, 0.97)
    glf_distance_98 = fit.getValueWithVerticalOffset(x_distance_glf, y_distance_glf,
                                                 popt_distance_glf, 0.98)
    glf_distance_99 = fit.getValueWithVerticalOffset(x_distance_glf, y_distance_glf,
                                                 popt_distance_glf, 0.99)
    plot.plotGLFFitDistance(frame_number_list_distance, distance_list, x_distance_glf, y_distance_glf,
                                saving_path + patient + "_" + sequence_number +
                                "_Generalized_Logistic_Function_Vocal_Fold_Edge_Distance.png")
    print("Vocal fold edge distance (generalized logistic function): ",
          [glf_distance_95, glf_distance_96, glf_distance_97, glf_distance_98, glf_distance_99])
    file.write("LAR onset frame via vocal fold edge distance "
               "(generalized logistic function, 95/96/97/98/99% decline): ")
    file.write(str([glf_distance_95, glf_distance_96, glf_distance_97, glf_distance_98, glf_distance_99]))
    file.write("\n")
    file.write("Parameters (generalized logistic function, vocal fold edge distance): ")
    file.write(str(popt_distance_glf))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_distance):
        error_sum += pow(abs(distance_list[i] - fit.generalized_logistic_function(frame_number_list_distance[i],
                                                                   *popt_distance_glf)), 2)
    error_sum /= last_index_fit_distance
    rmse = math.sqrt(error_sum)
    file.write("RMSE (generalized logistic function, vocal fold edge distance) in pixel: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=34).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_distance):
        error_sum_MAE += abs(distance_list[i] - fit.generalized_logistic_function(frame_number_list_distance[i],
                                                                   *popt_distance_glf))
    error_sum_MAE /= last_index_fit_distance
    file.write("MAE (generalized logistic function, vocal fold edge distance) in pixel: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=35).value = error_sum_MAE
except:
    file.write("Vocal fold edge distance: asymmetrical sigmoid fit with vertical offset not successful!\n\n")

# Gompertz-like function (distance)
try:
    x_distance_Gompertz, y_distance_Gompertz, popt_distance_gompertz = \
        fit.fit_gompertz(frame_number_list_distance, distance_list)
    gompertz_distance_95 = fit.getValueGompertz(x_distance_Gompertz, y_distance_Gompertz,
                                                 popt_distance_gompertz, 0.95)
    gompertz_distance_96 = fit.getValueGompertz(x_distance_Gompertz, y_distance_Gompertz,
                                                 popt_distance_gompertz, 0.96)
    gompertz_distance_97 = fit.getValueGompertz(x_distance_Gompertz, y_distance_Gompertz,
                                                 popt_distance_gompertz, 0.97)
    gompertz_distance_98 = fit.getValueGompertz(x_distance_Gompertz, y_distance_Gompertz,
                                                 popt_distance_gompertz, 0.98)
    gompertz_distance_99 = fit.getValueGompertz(x_distance_Gompertz, y_distance_Gompertz,
                                                 popt_distance_gompertz, 0.99)

    plot.plotGompertzFitDistance(frame_number_list_distance, distance_list, x_distance_Gompertz, y_distance_Gompertz,
                                saving_path + patient + "_" + sequence_number +
                                "_Gompertz_Vocal_Fold_Edge_Distance.png")

    print("Vocal fold edge distance (Gompertz-like fit): ",
          [gompertz_distance_95, gompertz_distance_96, gompertz_distance_97,
           gompertz_distance_98, gompertz_distance_99])

    file.write("LAR onset frame via vocal fold edge distance (Gompertz-like fit, 95/96/97/98/99% decline): ")
    file.write(str([gompertz_distance_95, gompertz_distance_96, gompertz_distance_97,
                       gompertz_distance_98, gompertz_distance_99]))
    file.write("\n")
    file.write("Parameters (Gompertz-like fit, vocal fold edge distance): ")
    file.write(str(popt_distance_gompertz))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_distance):
        error_sum += pow(abs(distance_list[i] - fit.gompertz(frame_number_list_distance[i],
                                                                   *popt_distance_gompertz)), 2)
    error_sum /= last_index_fit_distance
    rmse = math.sqrt(error_sum)
    file.write("RMSE (Gompertz-like function, vocal fold edge distance) in pixel: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=36).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_distance):
        error_sum_MAE += abs(distance_list[i] - fit.gompertz(frame_number_list_distance[i],
                                                                   *popt_distance_gompertz))
    error_sum_MAE /= last_index_fit_distance
    file.write("MAE (Gompertz-like function, vocal fold edge distance) in pixel: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=37).value = error_sum_MAE
except:
    file.write("Vocal fold edge distance: Gompertz-like fit not successful!\n\n")

# cubic polynomial (distance)
try:
    x_distance_cubic, y_distance_cubic, popt_distance_cubic = fit.fit_cubic(frame_number_list_distance, distance_list)

    cubic_distance_95 = fit.getValueCubic(x_distance_cubic, y_distance_cubic,
                                                popt_distance_cubic, 0.95)
    cubic_distance_96 = fit.getValueCubic(x_distance_cubic, y_distance_cubic,
                                                popt_distance_cubic, 0.96)
    cubic_distance_97 = fit.getValueCubic(x_distance_cubic, y_distance_cubic,
                                                popt_distance_cubic, 0.97)
    cubic_distance_98 = fit.getValueCubic(x_distance_cubic, y_distance_cubic,
                                                popt_distance_cubic, 0.98)
    cubic_distance_99 = fit.getValueCubic(x_distance_cubic, y_distance_cubic,
                                                popt_distance_cubic, 0.99)

    plot.plotCubicFitDistance(frame_number_list_distance, distance_list, x_distance_cubic, y_distance_cubic,
                             saving_path + patient + "_" + sequence_number +
                             "_Cubic_Vocal_Fold_Edge_Distance.png")

    print("Vocal fold edge distance (cubic): ", [cubic_distance_95, cubic_distance_96,
                              cubic_distance_97, cubic_distance_98,
                              cubic_distance_99])

    file.write("LAR onset frame via vocal fold edge distance (cubic fit function, 95/96/97/98/99% decline): ")
    file.write(str([cubic_distance_95, cubic_distance_96, cubic_distance_97, cubic_distance_98, cubic_distance_99]))
    file.write("\n")
    file.write("Parameters (cubic fit function, vocal fold edge distance): ")
    file.write(str(popt_distance_cubic))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_distance):
        error_sum += pow(abs(distance_list[i] - fit.cubic(frame_number_list_distance[i], *popt_distance_cubic)), 2)
    error_sum /= last_index_fit_distance
    rmse = math.sqrt(error_sum)
    file.write("RMSE (cubic fit function, vocal fold edge distance) in pixel: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=38).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_distance):
        error_sum_MAE += abs(distance_list[i] - fit.cubic(frame_number_list_distance[i], *popt_distance_cubic))
    error_sum_MAE /= last_index_fit_distance
    file.write("MAE (cubic fit function, vocal fold edge distance) in pixel: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=39).value = error_sum_MAE
except:
    file.write("Vocal fold edge distance: cubic polynomial fit not successful!\n\n")

# GLOTTAL ANGLE

file.write("GLOTTAL ANGLE\n\n")

# symmetrical sigmoid fit without offset (angle)
try:
    x_angle_sigmoid, y_angle_sigmoid, popt_angle_sigmoid = fit.sigmoid_fit(frame_number_list_angle, angle_list)

    sigmoid_angle_95 = fit.getValueNoVerticalOffset(x_angle_sigmoid, y_angle_sigmoid, popt_angle_sigmoid, 0.95)
    sigmoid_angle_96 = fit.getValueNoVerticalOffset(x_angle_sigmoid, y_angle_sigmoid, popt_angle_sigmoid, 0.96)
    sigmoid_angle_97 = fit.getValueNoVerticalOffset(x_angle_sigmoid, y_angle_sigmoid, popt_angle_sigmoid, 0.97)
    sigmoid_angle_98 = fit.getValueNoVerticalOffset(x_angle_sigmoid, y_angle_sigmoid, popt_angle_sigmoid, 0.98)
    sigmoid_angle_99 = fit.getValueNoVerticalOffset(x_angle_sigmoid, y_angle_sigmoid, popt_angle_sigmoid, 0.99)

    plot.plotSigmoidFitAngle(frame_number_list_angle, angle_list, x_angle_sigmoid, y_angle_sigmoid, saving_path +
                             patient + "_" + sequence_number + "_Symmetrical_Sigmoid_Glottal_Angle.png")

    print("Glottal angle (symm. sigmoid without vertical offset): ",
          [sigmoid_angle_95, sigmoid_angle_96, sigmoid_angle_97, sigmoid_angle_98, sigmoid_angle_99])

    file.write("LAR onset frame via glottal angle (symmetrical sigmoid, 95/96/97/98/99% decline): ")
    file.write(str([sigmoid_angle_95, sigmoid_angle_96, sigmoid_angle_97, sigmoid_angle_98, sigmoid_angle_99]))
    file.write("\n")
    file.write("Parameters (symmetrical sigmoid fit, glottal angle): ")
    file.write(str(popt_angle_sigmoid))
    file.write("\n")

    # calculate RMSE
    # index of last data point used for glottal angle fitting
    last_index_fit_angle = frame_number_list_angle.index(len(x_angle_sigmoid) - 1)
    error_sum = 0
    for i in range(0, last_index_fit_angle):
        error_sum += pow(abs(angle_list[i] - fit.sigmoid(frame_number_list_angle[i], *popt_angle_sigmoid)), 2)
    error_sum /= last_index_fit_angle
    rmse = math.sqrt(error_sum)
    file.write("RMSE (symmetrical sigmoid, glottal angle) in degree: ")
    file.write(str(rmse))
    file.write("\n")

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_angle):
        error_sum_MAE += abs(angle_list[i] - fit.sigmoid(frame_number_list_angle[i], *popt_angle_sigmoid))
    error_sum_MAE /= last_index_fit_angle
    file.write("MAE (symmetrical sigmoid, glottal angle) in degree: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # CALCULATION OF MEAN ANGULAR VELOCITY
    # (symmetrical sigmoid fit without vertical offset)

    # create x values; frame indices span 0 to x_angle_sigmoid[-1]
    # x_angle_sigmoid = np.linspace(0, 5000, 5001)
    x_angle_sigmoid_mean_calc = np.linspace(0, x_angle_sigmoid[-1], num=int(x_angle_sigmoid[-1] + 1))
    # calculate associated function values
    y_angle_sigmoid_mean_calc = fit.sigmoid(x_angle_sigmoid_mean_calc, *popt_angle_sigmoid)

    # if last element of "y_angle_sigmoid" smaller than (0.2*a) of sigmoid fit function
    if y_angle_sigmoid_mean_calc[-1] < popt_angle_sigmoid[0] * 0.2:
        sigmoid_angle_80 = fit.getValueNoVerticalOffset(x_angle_sigmoid_mean_calc, y_angle_sigmoid_mean_calc,
                                                        popt_angle_sigmoid, 0.8)
        sigmoid_angle_20 = fit.getValueNoVerticalOffset(x_angle_sigmoid_mean_calc, y_angle_sigmoid_mean_calc,
                                                        popt_angle_sigmoid, 0.2)
        print(sigmoid_angle_80)
        print(sigmoid_angle_20)
        # calculate mean angular velocity in degrees per frame
        mean_angular_velocity = (fit.sigmoid(sigmoid_angle_20, *popt_angle_sigmoid) -
                                 fit.sigmoid(sigmoid_angle_80, *popt_angle_sigmoid)) / (sigmoid_angle_20 -
                                                                                        sigmoid_angle_80)
        # convert into degrees per second using known frame rate of 4000 Hz
        mean_angular_velocity *= 4000.0
    else:
        mean_angular_velocity = None

    # CALCULATION OF MAXIMUM ANGULAR VELOCITY (ALWAYS POSSIBLE IF FIT SUCCESSFUL)
    max_angular_velocity_sigmoid_without_vert_offset = fit.getMaxAngVelocitySigmoidNoVertOffset(*popt_angle_sigmoid)
    # convert into degrees per second using known frame rate of 4000 Hz
    max_angular_velocity_sigmoid_without_vert_offset *= 4000.0

except:
    file.write("Glottal angle: symmetrical sigmoid fit without vertical offset not successful!\n\n")

# symmetrical sigmoid fit with vertical offset (angle)
try:
    x_angle_sigmoid_offset, y_angle_sigmoid_offset, popt_angle_sigmoid_offset = \
        fit.sigmoid_fit_offset(frame_number_list_angle, angle_list)
    sigmoid_angle_95_offset = fit.getValueWithVerticalOffset(x_angle_sigmoid_offset, y_angle_sigmoid_offset,
                                               popt_angle_sigmoid_offset, 0.95)
    sigmoid_angle_96_offset = fit.getValueWithVerticalOffset(x_angle_sigmoid_offset, y_angle_sigmoid_offset,
                                               popt_angle_sigmoid_offset, 0.96)
    sigmoid_angle_97_offset = fit.getValueWithVerticalOffset(x_angle_sigmoid_offset, y_angle_sigmoid_offset,
                                               popt_angle_sigmoid_offset, 0.97)
    sigmoid_angle_98_offset = fit.getValueWithVerticalOffset(x_angle_sigmoid_offset, y_angle_sigmoid_offset,
                                               popt_angle_sigmoid_offset, 0.98)
    sigmoid_angle_99_offset = fit.getValueWithVerticalOffset(x_angle_sigmoid_offset, y_angle_sigmoid_offset,
                                               popt_angle_sigmoid_offset, 0.99)

    plot.plotSigmoidFitAngleOffset(frame_number_list_angle, angle_list, x_angle_sigmoid_offset, y_angle_sigmoid_offset,
                             saving_path + patient + "_" + sequence_number +
                             "_Sigmoid_Fit_Vertical_Offset_Glottal_Angle.png")

    print("Glottal angle (symm. sigmoid with vertical offset): ",
          [sigmoid_angle_95_offset, sigmoid_angle_96_offset, sigmoid_angle_97_offset,
                              sigmoid_angle_98_offset, sigmoid_angle_99_offset])

    file.write("LAR onset frame via glottal angle "
               "(symmetrical sigmoid with vertical offset, 95/96/97/98/99% decline): ")
    file.write(str([sigmoid_angle_95_offset, sigmoid_angle_96_offset, sigmoid_angle_97_offset,
                    sigmoid_angle_98_offset, sigmoid_angle_99_offset]))
    file.write("\n")
    file.write("Parameters (symmetrical sigmoid fit with vertical offset, glottal angle): ")
    file.write(str(popt_angle_sigmoid_offset))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_angle):
        error_sum += pow(abs(angle_list[i] - fit.sigmoid_offset(frame_number_list_angle[i],
                                                                *popt_angle_sigmoid_offset)), 2)
    error_sum /= last_index_fit_angle
    rmse = math.sqrt(error_sum)
    file.write("RMSE (symmetrical sigmoid with vertical offset, glottal angle) in degree: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=40).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_angle):
        error_sum_MAE += abs(angle_list[i] - fit.sigmoid_offset(frame_number_list_angle[i],
                                                                *popt_angle_sigmoid_offset))
    error_sum_MAE /= last_index_fit_angle
    file.write("MAE (symmetrical sigmoid with vertical offset, glottal angle) in degree: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=41).value = error_sum_MAE

    # CALCULATION OF MEAN ANGULAR VELOCITY
    # (symmetrical sigmoid fit with vertical offset)

    # create x values; frame indices span 0 to x_angle_sigmoid_offset[-1]
    # x_angle_sigmoid_offset = np.linspace(0, 5000, 5001)
    x_angle_sigmoid_offset_mean_calc = np.linspace(0, x_angle_sigmoid_offset[-1],
                                                   num=int(x_angle_sigmoid_offset[-1] + 1))
    # calculate associated function values
    y_angle_sigmoid_offset_mean_calc = fit.sigmoid_offset(x_angle_sigmoid_offset_mean_calc, *popt_angle_sigmoid_offset)

    if y_angle_sigmoid_offset_mean_calc[-1] < (popt_angle_sigmoid_offset[0] * 0.2) + popt_angle_sigmoid_offset[3]:
        sigmoid_angle_80_offset = fit.getValueWithVerticalOffset(x_angle_sigmoid_offset_mean_calc,
                                                                 y_angle_sigmoid_offset_mean_calc,
                                                                 popt_angle_sigmoid_offset, 0.8)
        sigmoid_angle_20_offset = fit.getValueWithVerticalOffset(x_angle_sigmoid_offset_mean_calc,
                                                                 y_angle_sigmoid_offset_mean_calc,
                                                                 popt_angle_sigmoid_offset, 0.2)
        print(sigmoid_angle_80_offset)
        print(sigmoid_angle_20_offset)
        # calculate mean angular velocity in degrees per frame
        mean_angular_velocity_offset = (fit.sigmoid_offset(sigmoid_angle_20_offset, *popt_angle_sigmoid_offset) -
                                        fit.sigmoid_offset(sigmoid_angle_80_offset, *popt_angle_sigmoid_offset)) / \
                                       (sigmoid_angle_20_offset - sigmoid_angle_80_offset)
        # convert into degrees per second using known frame rate of 4000 Hz
        mean_angular_velocity_offset *= 4000.0
    else:
        mean_angular_velocity_offset = None

    # CALCULATION OF MAXIMUM ANGULAR VELOCITY (ALWAYS POSSIBLE IF FIT SUCCESSFUL)
    max_angular_velocity_sigmoid_offset = fit.getMaxAngVelocitySigmoidWithVertOffset(*popt_angle_sigmoid_offset)
    # convert into degrees per second using known frame rate of 4000 Hz
    max_angular_velocity_sigmoid_offset *= 4000.0
except:
    file.write("Glottal angle: symmetrical sigmoid fit with vertical offset not successful!\n\n")

# generalized logistic function (angle)
try:
    x_angle_glf, y_angle_glf, popt_angle_glf = \
        fit.fit_generalized_logistic_function(frame_number_list_angle, angle_list)

    glf_angle_95 = fit.getValueWithVerticalOffset(x_angle_glf, y_angle_glf,
                                                     popt_angle_glf, 0.95)
    glf_angle_96 = fit.getValueWithVerticalOffset(x_angle_glf, y_angle_glf,
                                                     popt_angle_glf, 0.96)
    glf_angle_97 = fit.getValueWithVerticalOffset(x_angle_glf, y_angle_glf,
                                                     popt_angle_glf, 0.97)
    glf_angle_98 = fit.getValueWithVerticalOffset(x_angle_glf, y_angle_glf,
                                                     popt_angle_glf, 0.98)
    glf_angle_99 = fit.getValueWithVerticalOffset(x_angle_glf, y_angle_glf,
                                                     popt_angle_glf, 0.99)

    plot.plotGLFFitAngle(frame_number_list_angle, angle_list, x_angle_glf, y_angle_glf,
                             saving_path + patient + "_" + sequence_number +
                             "_Generalized_Logistic_Function_Glottal_Angle.png")

    print("Glottal angle (generalized logistic function): ", [glf_angle_95, glf_angle_96, glf_angle_97, glf_angle_98,
                                                              glf_angle_99])

    file.write("LAR onset frame via glottal angle (generalized logistic function, 95/96/97/98/99% decline): ")
    file.write(str([glf_angle_95, glf_angle_96, glf_angle_97, glf_angle_98, glf_angle_99]))
    file.write("\n")
    file.write("Parameters (generalized logistic function, glottal angle): ")
    file.write(str(popt_angle_glf))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_angle):
        error_sum += pow(abs(angle_list[i] - fit.generalized_logistic_function(frame_number_list_angle[i],
                                                                *popt_angle_glf)), 2)
    error_sum /= last_index_fit_angle
    rmse = math.sqrt(error_sum)
    file.write("RMSE (generalized logistic function, glottal angle) in degree: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=42).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_angle):
        error_sum_MAE += abs(angle_list[i] - fit.generalized_logistic_function(frame_number_list_angle[i],
                                                                *popt_angle_glf))
    error_sum_MAE /= last_index_fit_angle
    file.write("MAE (generalized logistic function, glottal angle) in degree: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=43).value = error_sum_MAE

    # CALCULATION OF MEAN ANGULAR VELOCITY
    # (generalized logistic function)

    # create x values; indices of frames used for fitting span 0 to x_angle_glf[-1]
    x_angle_glf_mean_calc = np.linspace(0, x_angle_glf[-1], num=int(x_angle_glf[-1] + 1))
    # calculate associated function values
    y_angle_glf_mean_calc = fit.generalized_logistic_function(x_angle_glf_mean_calc, *popt_angle_glf)

    # if angle decreases to below 20% of function span (a)
    if y_angle_glf_mean_calc[-1] < (popt_angle_glf[0] * 0.2) + popt_angle_glf[3]:
        angle_80_generalized_logistic_function = fit.getValueWithVerticalOffset(
            x_angle_glf_mean_calc, y_angle_glf_mean_calc, popt_angle_glf, 0.8)
        angle_20_generalized_logistic_function = fit.getValueWithVerticalOffset(
            x_angle_glf_mean_calc, y_angle_glf_mean_calc, popt_angle_glf, 0.2)
        print(angle_80_generalized_logistic_function)
        print(angle_20_generalized_logistic_function)
        # declare mean angular velocity in degrees per frame
        mean_angular_velocity_glf = (fit.generalized_logistic_function(angle_20_generalized_logistic_function,
                                                                       *popt_angle_glf) -
                                     fit.generalized_logistic_function(angle_80_generalized_logistic_function,
                                                                       *popt_angle_glf)) / \
                                    (angle_20_generalized_logistic_function - angle_80_generalized_logistic_function)

        # convert into degrees per second using known frame rate of 4000 Hz
        mean_angular_velocity_glf *= 4000.0
    else:
        mean_angular_velocity_glf = None

    # CALCULATION OF MAXIMUM ANGULAR VELOCITY (ALWAYS POSSIBLE IF FIT SUCCESSFUL)
    max_angular_velocity_glf = fit.getMaxAngVelocityGeneralizedLogisticFunction(*popt_angle_glf)
    # convert into degrees per second using known frame rate of 4000 Hz
    max_angular_velocity_glf *= 4000.0
except:
    file.write("Glottal angle: generalized logistic function fit not successful!\n\n")

# Gompertz-like function (angle)
try:
    x_angle_gompertz, y_angle_gompertz, popt_angle_gompertz = fit.fit_gompertz(frame_number_list_angle, angle_list)

    gompertz_angle_95 = fit.getValueGompertz(x_angle_gompertz, y_angle_gompertz,
                                                popt_angle_gompertz, 0.95)
    gompertz_angle_96 = fit.getValueGompertz(x_angle_gompertz, y_angle_gompertz,
                                                popt_angle_gompertz, 0.96)
    gompertz_angle_97 = fit.getValueGompertz(x_angle_gompertz, y_angle_gompertz,
                                                popt_angle_gompertz, 0.97)
    gompertz_angle_98 = fit.getValueGompertz(x_angle_gompertz, y_angle_gompertz,
                                                popt_angle_gompertz, 0.98)
    gompertz_angle_99 = fit.getValueGompertz(x_angle_gompertz, y_angle_gompertz,
                                                popt_angle_gompertz, 0.99)

    plot.plotGompertzFitAngle(frame_number_list_angle, angle_list, x_angle_gompertz, y_angle_gompertz,
                             saving_path + patient + "_" + sequence_number +
                             "_Gompertz_Glottal_Angle.png")

    print("Glottal angle (Gompertz): ", [gompertz_angle_95, gompertz_angle_96,
                              gompertz_angle_97, gompertz_angle_98,
                              gompertz_angle_99])

    file.write("LAR onset frame via glottal angle (Gompertz-like function, 95/96/97/98/99% decline): ")
    file.write(str([gompertz_angle_95, gompertz_angle_96, gompertz_angle_97, gompertz_angle_98, gompertz_angle_99]))
    file.write("\n")
    file.write("Parameters (Gompertz-like function, glottal angle): ")
    file.write(str(popt_angle_gompertz))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_angle):
        error_sum += pow(abs(angle_list[i] - fit.gompertz(frame_number_list_angle[i],
                                                                *popt_angle_gompertz)), 2)
    error_sum /= last_index_fit_angle
    rmse = math.sqrt(error_sum)
    file.write("RMSE (Gompertz-like function, glottal angle) in degree: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=44).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_angle):
        error_sum_MAE += abs(angle_list[i] - fit.gompertz(frame_number_list_angle[i],
                                                                *popt_angle_gompertz))
    error_sum_MAE /= last_index_fit_angle
    file.write("MAE (Gompertz-like function, glottal angle) in degree: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=45).value = error_sum_MAE

    # CALCULATION OF MEAN ANGULAR VELOCITY
    # (Gompertz-like function)

    # x_angle_gompertz = np.linspace(0, 5000, 5001)

    # create x values; indices of frames used for fitting span 0 to x_angle_gompertz[-1]
    x_angle_gompertz_mean_calc = np.linspace(0, x_angle_gompertz[-1], num=int(x_angle_gompertz[-1] + 1))
    # calculate associated function values
    y_angle_gompertz_mean_calc = fit.gompertz(x_angle_gompertz_mean_calc, *popt_angle_gompertz)

    # if angle decreases to below 20% of function span (d-a)
    if y_angle_gompertz_mean_calc[-1] < (popt_angle_gompertz[3]-popt_angle_gompertz[0]) * 0.2 + popt_angle_gompertz[0]:
        angle_80_gompertz = fit.getValueGompertz(x_angle_gompertz_mean_calc, y_angle_gompertz_mean_calc,
                                                                 popt_angle_gompertz, 0.8)
        angle_20_gompertz = fit.getValueGompertz(x_angle_gompertz_mean_calc, y_angle_gompertz_mean_calc,
                                                                 popt_angle_gompertz, 0.2)
        print(angle_80_gompertz)
        print(angle_20_gompertz)
        # declare mean angular velocity in degrees per frame
        mean_angular_velocity_gompertz = (fit.gompertz(angle_20_gompertz, *popt_angle_gompertz) -
                                              fit.gompertz(angle_80_gompertz, *popt_angle_gompertz)) / \
                                             (angle_20_gompertz - angle_80_gompertz)
        # convert into degrees per second using known frame rate of 4000 Hz
        mean_angular_velocity_gompertz *= 4000.0
    else:
        mean_angular_velocity_gompertz = None

    # CALCULATION OF MAXIMUM ANGULAR VELOCITY (ALWAYS POSSIBLE IF FIT SUCCESSFUL)
    max_angular_velocity_gompertz = fit.getMaxAngVelocityGompertz(*popt_angle_gompertz)
    # convert into degrees per second using known frame rate of 4000 Hz
    max_angular_velocity_gompertz *= 4000.0
except:
    file.write("Glottal angle: Gompertz-like function fit not successful!\n\n")

# cubic function (angle)
try:
    # obtain frame indices and associated values of cubic fit function up to automatically detected stop frame
    x_angle_cubic, y_angle_cubic, popt_angle_cubic = fit.fit_cubic(frame_number_list_angle, angle_list)

    cubic_angle_95 = fit.getValueCubic(x_angle_cubic, y_angle_cubic,
                                                popt_angle_cubic, 0.95)
    cubic_angle_96 = fit.getValueCubic(x_angle_cubic, y_angle_cubic,
                                                popt_angle_cubic, 0.96)
    cubic_angle_97 = fit.getValueCubic(x_angle_cubic, y_angle_cubic,
                                                popt_angle_cubic, 0.97)
    cubic_angle_98 = fit.getValueCubic(x_angle_cubic, y_angle_cubic,
                                                popt_angle_cubic, 0.98)
    cubic_angle_99 = fit.getValueCubic(x_angle_cubic, y_angle_cubic,
                                                popt_angle_cubic, 0.99)

    plot.plotCubicFitAngle(frame_number_list_angle, angle_list, x_angle_cubic, y_angle_cubic,
                             saving_path + patient + "_" + sequence_number +
                             "_Cubic_Glottal_Angle.png")

    print("Glottal angle (cubic): ", [cubic_angle_95, cubic_angle_96, cubic_angle_97, cubic_angle_98, cubic_angle_99])

    file.write("LAR onset frame via glottal angle (cubic fit function, 95/96/97/98/99% decline): ")
    file.write(str([cubic_angle_95, cubic_angle_96, cubic_angle_97, cubic_angle_98, cubic_angle_99]))
    file.write("\n")
    file.write("Parameters (cubic fit function, glottal angle): ")
    file.write(str(popt_angle_cubic))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_angle):
        error_sum += pow(abs(angle_list[i] - fit.cubic(frame_number_list_angle[i], *popt_angle_cubic)), 2)
    error_sum /= last_index_fit_angle
    rmse = math.sqrt(error_sum)
    file.write("RMSE (cubic fit function, glottal angle) in degree: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=46).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_angle):
        error_sum_MAE += abs(angle_list[i] - fit.cubic(frame_number_list_angle[i], *popt_angle_cubic))
    error_sum_MAE /= last_index_fit_angle
    file.write("MAE (cubic fit function, glottal angle) in degree: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=47).value = error_sum_MAE

    # CALCULATION OF MEAN ANGULAR VELOCITY
    # (cubic fit function)

    # create x values; indices of frames used for fitting span 0 to x_angle_cubic[-1]
    x_angle_cubic_mean_calc = np.linspace(0, x_angle_cubic[-1], num=int(x_angle_cubic[-1] + 1))
    # calculate associated function values
    y_angle_cubic_mean_calc = fit.cubic(x_angle_cubic_mean_calc, *popt_angle_cubic)

    # determine roots of derivative of cubic fit function
    roots = poly.polyroots([popt_angle_cubic[2], 2*popt_angle_cubic[1], 3*popt_angle_cubic[0]])
    file.write("Roots of derivative of cubic: ")
    file.write(str(roots[0]))
    file.write(", ")
    file.write(str(roots[1]))
    file.write("\n\n")

    # if angle decreases to below 20% of function value at rightmost stationary point of cubic fit function
    if y_angle_cubic_mean_calc[-1] < (fit.cubic(np.max(roots), *popt_angle_cubic) * 0.2):
        # determine time stamps of values of interest on cubic fit function
        angle_80_cubic = fit.getValueCubic(x_angle_cubic_mean_calc, y_angle_cubic_mean_calc, popt_angle_cubic, 0.8)
        angle_20_cubic = fit.getValueCubic(x_angle_cubic_mean_calc, y_angle_cubic_mean_calc, popt_angle_cubic, 0.2)
        print(angle_80_cubic)
        print(angle_20_cubic)
        # declare mean angular velocity in degrees per frame
        mean_angular_velocity_cubic = (fit.cubic(angle_20_cubic, *popt_angle_cubic) -
                                       fit.cubic(angle_80_cubic, *popt_angle_cubic)) / (angle_20_cubic - angle_80_cubic)
        # convert into degrees per second using known frame rate of 4000 Hz
        mean_angular_velocity_cubic *= 4000.0
    else:
        file.write("Calculation of mean angular velocity with cubic fit not successful "
                   "(function value not below 20%)!\n\n")
        mean_angular_velocity_cubic = None
except:
    file.write("Glottal angle: cubic polynomial fit not successful!\n\n")

# plot comparison of fit functions for glottal angle
# plot.plotTwoFitsAngle(frame_number_list_angle, angle_list, x_angle_sigmoid, y_angle_sigmoid,
#                  x_angle_gompertz, y_angle_gompertz, saving_path + patient + "_" + sequence_number +
#                  "_Sigmoid_and_Gompertz_Glottal_Angle.png")

# GLOTTAL AREA

file.write("GLOTTAL AREA\n\n")

# symmetrical sigmoid fit with vertical offset (area)
try:
    # was "0:len(frame_number_list_area)-3]"
    x_area_sigmoid_offset, y_area_sigmoid_offset, popt_area_sigmoid_offset = \
        fit.sigmoid_fit_offset(frame_number_list_area[0:len(frame_number_list_area)],
                               area_list[0:len(frame_number_list_area)])
    sigmoid_area_95 = fit.getValueWithVerticalOffset(x_area_sigmoid_offset, y_area_sigmoid_offset,
                                                     popt_area_sigmoid_offset, 0.95)
    sigmoid_area_96 = fit.getValueWithVerticalOffset(x_area_sigmoid_offset, y_area_sigmoid_offset,
                                                     popt_area_sigmoid_offset, 0.96)
    sigmoid_area_97 = fit.getValueWithVerticalOffset(x_area_sigmoid_offset, y_area_sigmoid_offset,
                                                     popt_area_sigmoid_offset, 0.97)
    sigmoid_area_98 = fit.getValueWithVerticalOffset(x_area_sigmoid_offset, y_area_sigmoid_offset,
                                                     popt_area_sigmoid_offset, 0.98)
    sigmoid_area_99 = fit.getValueWithVerticalOffset(x_area_sigmoid_offset, y_area_sigmoid_offset,
                                                     popt_area_sigmoid_offset, 0.99)

    plot.plotSigmoidFitAreaOffset(frame_number_list_area, area_list, x_area_sigmoid_offset, y_area_sigmoid_offset,
                                  saving_path + patient + "_" + sequence_number +
                                  "_Sigmoid_Fit_Vertical_Offset_Glottal_Area.png")

    print("Glottal area (symmetrical sigmoid fit with vertical offset): ", [sigmoid_area_95, sigmoid_area_96,
                                                                            sigmoid_area_97, sigmoid_area_98,
                                                                            sigmoid_area_99])

    file.write("LAR onset frame via glottal area (symmetrical sigmoid fit with vertical offset, "
               "95/96/97/98/99% decline): ")
    file.write(str([sigmoid_area_95, sigmoid_area_96, sigmoid_area_97, sigmoid_area_98, sigmoid_area_99]))
    file.write("\n")

    file.write("Parameters (symmetrical sigmoid fit with vertical offset, glottal area): ")
    file.write(str(popt_area_sigmoid_offset))
    file.write("\n")

    # calculate RMSE
    # index of last data point used for glottal area fitting
    last_index_fit_area = frame_number_list_area.index(len(x_area_sigmoid_offset) - 1)
    error_sum = 0
    for i in range(0, last_index_fit_area):
        error_sum += pow(abs(area_list[i] - fit.sigmoid_offset(frame_number_list_area[i],
                                                               *popt_area_sigmoid_offset)), 2)
    error_sum /= last_index_fit_area
    rmse = math.sqrt(error_sum)
    file.write("RMSE (symmetrical sigmoid fit with vertical offset, glottal area) in percent: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=48).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_area):
        error_sum_MAE += abs(area_list[i] - fit.sigmoid_offset(frame_number_list_area[i], *popt_area_sigmoid_offset))
    error_sum_MAE /= last_index_fit_area
    file.write("MAE (symmetrical sigmoid fit with vertical offset, glottal area) in percent: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=49).value = error_sum_MAE

    # check if vertical offset parameter value below 0.1
    if popt_area_sigmoid_offset[3] < 0.1:
        file.write("Glottal closure completed (symmetrical sigmoid with vertical offset): yes\n\n")
    else:
        file.write("Glottal closure completed (symmetrical sigmoid with vertical offset): no\n\n")
except:
    file.write("Glottal area: symmetrical sigmoid fit with vertical offset not successful!\n\n")

# generalized logistic function (area)
try:
    # was "0:len(frame_number_list_area)-3]"
    x_area_glf, y_area_glf, popt_area_glf = \
        fit.fit_generalized_logistic_function(frame_number_list_area[0:len(frame_number_list_area)],
                                            area_list[0:len(frame_number_list_area)])
    glf_area_95 = fit.getValueWithVerticalOffset(x_area_glf, y_area_glf, popt_area_glf, 0.95)
    glf_area_96 = fit.getValueWithVerticalOffset(x_area_glf, y_area_glf, popt_area_glf, 0.96)
    glf_area_97 = fit.getValueWithVerticalOffset(x_area_glf, y_area_glf, popt_area_glf, 0.97)
    glf_area_98 = fit.getValueWithVerticalOffset(x_area_glf, y_area_glf, popt_area_glf, 0.98)
    glf_area_99 = fit.getValueWithVerticalOffset(x_area_glf, y_area_glf, popt_area_glf, 0.99)

    plot.plotGLFFitArea(frame_number_list_area, area_list, x_area_glf, y_area_glf, saving_path + patient +
                            "_" + sequence_number + "_Generalized_Logistic_Function_Glottal_Area.png")

    print("Glottal area (generalized logistic function): ", [glf_area_95, glf_area_96, glf_area_97,
                                                             glf_area_98, glf_area_99])

    file.write("Glottal area (generalized logistic function, 95/96/97/98/99% decline)): ")
    file.write(str([glf_area_95, glf_area_96, glf_area_97, glf_area_98, glf_area_99]))
    file.write("\n")
    file.write("Parameters (generalized logistic function, glottal area): ")
    file.write(str(popt_area_glf))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_area):
        error_sum += pow(abs(area_list[i] - fit.generalized_logistic_function(frame_number_list_area[i],
                                                               *popt_area_glf)), 2)
    error_sum /= last_index_fit_area
    rmse = math.sqrt(error_sum)
    file.write("RMSE (generalized logistic function, glottal area) in percent: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=50).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_area):
        error_sum_MAE += abs(area_list[i] - fit.generalized_logistic_function(frame_number_list_area[i],
                                                                              *popt_area_glf))
    error_sum_MAE /= last_index_fit_area
    file.write("MAE (generalized logistic function, glottal area) in percent: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=51).value = error_sum_MAE

    # if vertical offset of glottal area below 0.1
    if popt_area_glf[3] < 0.1:
        file.write("Glottal closure completed (generalized logistic function): yes\n\n")
    else:
        file.write("Glottal closure completed (generalized logistic function): no\n\n")
except:
    file.write("Glottal area: generalized logistic function fit not successful!\n\n")

# Gompertz-like function (area)
try:
    # was "0:len(frame_number_list_area)-3]"
    x_area_gompertz, y_area_gompertz, popt_area_gompertz = \
        fit.fit_gompertz(frame_number_list_area[0:len(frame_number_list_area)],
                                              area_list[0:len(frame_number_list_area)])
    gompertz_area_95 = fit.getValueGompertz(x_area_gompertz, y_area_gompertz, popt_area_gompertz, 0.95)
    gompertz_area_96 = fit.getValueGompertz(x_area_gompertz, y_area_gompertz, popt_area_gompertz, 0.96)
    gompertz_area_97 = fit.getValueGompertz(x_area_gompertz, y_area_gompertz, popt_area_gompertz, 0.97)
    gompertz_area_98 = fit.getValueGompertz(x_area_gompertz, y_area_gompertz, popt_area_gompertz, 0.98)
    gompertz_area_99 = fit.getValueGompertz(x_area_gompertz, y_area_gompertz, popt_area_gompertz, 0.99)

    plot.plotGompertzFitArea(frame_number_list_area, area_list,
                             x_area_gompertz, y_area_gompertz, saving_path + patient +
                             "_" + sequence_number + "_Gompertz_Glottal_Area.png")

    print("Glottal area (Gompertz-like function): ", [gompertz_area_95, gompertz_area_96,
                                                             gompertz_area_97, gompertz_area_98,
                                                             gompertz_area_99])

    file.write("Glottal area (Gompertz-like function, 95/96/97/98/99% decline)): ")
    file.write(str([gompertz_area_95, gompertz_area_96, gompertz_area_97, gompertz_area_98, gompertz_area_99]))
    file.write("\n")

    file.write("Parameters (Gompertz-like function, glottal area): ")
    file.write(str(popt_area_gompertz))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_area):
        error_sum += pow(abs(area_list[i] - fit.gompertz(frame_number_list_area[i], *popt_area_gompertz)), 2)
    error_sum /= last_index_fit_area
    rmse = math.sqrt(error_sum)
    file.write("RMSE (Gompertz-like function, glottal area) in percent: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=52).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_area):
        error_sum_MAE += abs(area_list[i] - fit.gompertz(frame_number_list_area[i], *popt_area_gompertz))
    error_sum_MAE /= last_index_fit_area
    file.write("MAE (Gompertz-like function, glottal area) in percent: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=53).value = error_sum_MAE

    # if vertical offset of glottal area below 0.1
    if popt_area_gompertz[0] < 0.1:
        file.write("Glottal closure completed (Gompertz-like function): yes\n\n")
    else:
        file.write("Glottal closure completed (Gompertz-like function): no\n\n")
except:
    file.write("Glottal area: Gompertz-like function fit not successful!\n\n")

# cubic polynomial (area)
try:
    # was "0:len(frame_number_list_area)-3]"
    x_area_cubic, y_area_cubic, popt_area_cubic = fit.fit_cubic(frame_number_list_area[0:len(frame_number_list_area)],
                                              area_list[0:len(frame_number_list_area)])

    cubic_area_95 = fit.getValueCubic(x_area_cubic, y_area_cubic, popt_area_cubic, 0.95)
    cubic_area_96 = fit.getValueCubic(x_area_cubic, y_area_cubic, popt_area_cubic, 0.96)
    cubic_area_97 = fit.getValueCubic(x_area_cubic, y_area_cubic, popt_area_cubic, 0.97)
    cubic_area_98 = fit.getValueCubic(x_area_cubic, y_area_cubic, popt_area_cubic, 0.98)
    cubic_area_99 = fit.getValueCubic(x_area_cubic, y_area_cubic, popt_area_cubic, 0.99)

    plot.plotCubicFitArea(frame_number_list_area, area_list, x_area_cubic, y_area_cubic,
                             saving_path + patient + "_" + sequence_number +
                             "_Cubic_Glottal_Area.png")

    print("Glottal area (cubic): ", [cubic_area_95, cubic_area_96, cubic_area_97, cubic_area_98, cubic_area_99])

    file.write("Glottal area (cubic fit function, 95/96/97/98/99% decline): ")
    file.write(str([cubic_area_95, cubic_area_96, cubic_area_97, cubic_area_98, cubic_area_99]))
    file.write("\n")
    file.write("Parameters (cubic fit function, glottal area): ")
    file.write(str(popt_area_cubic))
    file.write("\n")

    # calculate RMSE
    error_sum = 0
    for i in range(0, last_index_fit_area):
        error_sum += pow(abs(area_list[i] - fit.cubic(frame_number_list_area[i], *popt_area_cubic)), 2)
    error_sum /= last_index_fit_area
    rmse = math.sqrt(error_sum)
    file.write("RMSE (cubic fit function, glottal area) in percent: ")
    file.write(str(rmse))
    file.write("\n")

    # write RMSE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=54).value = rmse

    # calculate MAE
    error_sum_MAE = 0
    for i in range(0, last_index_fit_area):
        error_sum_MAE += abs(area_list[i] - fit.cubic(frame_number_list_area[i], *popt_area_cubic))
    error_sum_MAE /= last_index_fit_area
    file.write("MAE (cubic fit function, glottal area) in percent: ")
    file.write(str(error_sum_MAE))
    file.write("\n\n")

    # write MAE value to spreadsheet
    sheet.cell(row=spreadsheet_row, column=55).value = error_sum_MAE
except:
    file.write("Glottal area: cubic polynomial fit not successful!\n\n")

# save spreadsheet file
spreadsheet.save(filename=spreadsheet_path)

# LINEAR FIT (NOT USED FOR EVALUATION)
try:
    # (distance)
    if len(distance_list) > 7:
        impact_distance, lines_lar_begin_distance = fit.linear_fit(frame_number_list_distance, distance_list)

        plot.linearFitDistance(frame_number_list_distance, distance_list, lines_lar_begin_distance, impact_distance,
                               saving_path + patient + "_" + sequence_number +
                               "_Linear_Fit_Vocal_Fold_Edge_Distance.png")

        print("Vocal fold edge distance: ", impact_distance[0])

        file.write("LAR ONSET TIME (LINEAR FIT, NOT USED FOR EVALUATION)\n\n")

        file.write("LAR onset frame via vocal fold edge distance (linear fit, not used for evaluation): ")
        file.write(str(impact_distance[0]))
        file.write("\n")

    # (angle)
    if len(angle_list) > 7:
        impact_angle, lines_lar_begin_angle = fit.linear_fit(frame_number_list_angle, angle_list)

        plot.linearFitAngle(frame_number_list_angle, angle_list, lines_lar_begin_angle, impact_angle, saving_path +
                            patient + "_" + sequence_number + "_Linear_Fit_Glottal_Angle.png")

        print("Glottal angle: ", impact_angle[0])

        file.write("LAR onset frame via glottal angle (linear fit, not used for evaluation): ")
        file.write(str(impact_angle[0]))
        file.write("\n")

    # (area)
    if len(area_list) > 7:
        impact_area, lines_lar_begin_area = fit.linear_fit(frame_number_list_area, area_list)

        plot.linearFitArea(frame_number_list_area, area_list, lines_lar_begin_area, impact_area, saving_path +
                           patient + "_" + sequence_number + "_Linear_Fit_Glottal_Area.png")

        print("Glottal area: ", impact_area[0])

        file.write("LAR onset frame via glottal area (linear fit, not used for evaluation): ")
        file.write(str(impact_area[0]))
        file.write("\n\n")
except:
    pass

try:
    file.write("Initial glottal angle in degrees (first data point): " + str(angle_list[0]) + "\n\n")

    # if initial glottal angle below 20 degrees or initial vocal fold distance below 10 px: glottis pre-closed!
    if angle_list[0] < 20 or distance_list[0] < 10:
        file.write("Pre-closed glottis: yes\n\n")
    else:
        file.write("Pre-closed glottis: no\n\n")

    file.write("Avg. angular velocity of adduction (sigmoid without offset) in degrees/s: " +
               str(mean_angular_velocity) + "\n")
    file.write("Avg. angular velocity of adduction (sigmoid with offset) in degrees/s: " +
               str(mean_angular_velocity_offset) + "\n")
    file.write("Avg. angular velocity of adduction (generalized logistic function) in degrees/s: " +
               str(mean_angular_velocity_glf) + "\n")
    file.write("Avg. angular velocity of adduction (Gompertz-like function) in degrees/s: " +
               str(mean_angular_velocity_gompertz) + "\n")
    file.write("Avg. angular velocity of adduction (cubic fit function) in degrees/s: " +
               str(mean_angular_velocity_cubic) + "\n\n")

    file.write("Max. angular velocity of adduction (sigmoid without offset) in degrees/s: " +
               str(max_angular_velocity_sigmoid_without_vert_offset) + "\n")
    file.write("Max. angular velocity of adduction (sigmoid with offset) in degrees/s: " +
               str(max_angular_velocity_sigmoid_offset) + "\n")
    file.write("Max. angular velocity of adduction (generalized logistic function) in degrees/s: " +
               str(max_angular_velocity_glf) + "\n")
    file.write("Max. angular velocity of adduction (Gompertz-like function) in degrees/s: " +
               str(max_angular_velocity_gompertz) + "\n\n")

    # check if vertical offset parameter value below 0.1
    if popt_area_sigmoid_offset[3] < 0.1:
        file.write("Minimum relative glottal area with respect to initial area in percent: " +
                   str((np.min(area_list)/area_list[0])*100) + "\n\n")
        file.write("Duration of adduction phase in ms (vocal fold edge distance with offset - 98): ")
        file.write(str((frame_number_list_area[np.argmin(area_list)] - sigmoid_distance_98) / 4.0))
        file.write("\n")
    else:
        file.write("Relative glottal area with respect to initial area in percent: " +
                   str((np.min(area_list)/area_list[0]) * 100) + "\n")
except:
    pass

# close all windows
cv2.destroyAllWindows()

# list for inverse video playback
# remove last frame from list
# frame_list.pop()
# reverse order of (original) frames in list
frame_list.reverse()
# set last_frame_number to last frame in sequence (before application of .pop())
last_frame_number = frame_number

# identify instant with closed glottis/minimum glottis opening in ms
# (last element of frame index list of glottal area evaluation / frame rate in Hz) * 1000
glottal_closure = (frame_number_list_area[-1])/4

# show first (original) frame of inverted sequence
frame = frame_list[0]

if rotationCorrection:
    # rotate frame to correct glottal orientation
    frame = segmentation.rotate_frame(frame, -angle_glottal_midline)

frame_large = cv2.resize(frame, (int(2.0 * frame.shape[1]), int(2.0 * frame.shape[0])),
                                       interpolation=cv2.INTER_LINEAR)
cv2.imshow("Reverse Analysis", frame_large)
cv2.waitKey()
input_user = input("Is the glottis open and was glottal closure achieved during LAR? (y/n)\n")
if input_user == "y" and not (popt_area_sigmoid_offset[3] > 0.1):
    file.write("Glottal closure achieved and glottis open at end of sequence: yes\n\n")
    frame_number = 1

    # apply watershed segmentation with identified hysteresis threshold values on last grayscale frame of sequence
    label = segmentation.getLabels(frame_gray, low_canny_thresh, high_canny_thresh)

    watershed = segmentation.watershed_segmentation(cv2.cvtColor(frame_gray, cv2.COLOR_GRAY2BGR), label)

    # identify segment of watershed transformation result containing reference point
    if rotationCorrection:
        index_reference_point = segmentation.segment_glottis_point(watershed, ref_rot[0], ref_rot[1],
                                                                   frame_gray.shape[0], frame_gray.shape[1])
    else:
        index_reference_point = segmentation.segment_glottis_point(watershed, x, y,
                                                                   frame_gray.shape[0], frame_gray.shape[1])

    glottis_contour_watershed = segmentation.getGlottisContourFromWatershed(frame_gray.shape[0], frame_gray.shape[1],
                                                                            watershed, index_reference_point)

    # region growing (automated)
    extLeft, extRight, extTop, extBot = segmentation.getExtremePointsContour(glottis_contour_watershed)

    mask_grid_region_growing = segmentation.getGridForRegionGrowing(frame_gray.shape[0], frame_gray.shape[1],
                                                                    extLeft, extRight, extTop, extBot)

    seed_points = segmentation.getSeedPointsRegionGrowingFirstFrame(frame_gray.shape[0], frame_gray.shape[1],
                                                                    glottis_contour_watershed, mask_grid_region_growing)

    # # draw seed points
    # for seed in seed_points:
    #     frame = cv2.circle(frame, (int(seed[0]), int(seed[1])), 1, [255, 255, 0], -1)
    #
    # # draw contour from watershed
    # frame = cv2.drawContours(frame, [glottis_contour_watershed], 0, [255, 255, 0], 1)

    # calculate homogeneity criterion for region growing procedure
    homogeneity_criterion = segmentation.getHomogeneityCriterion(frame_gray, seed_points)
    # apply region growing procedure
    region_growing = segmentation.regionGrowing(frame_gray, seed_points, homogeneity_criterion)
    # obtain glottis contour from region growing result
    glottis_contour_region_growing = segmentation.getGlottisContourRegionGrowing(region_growing)

    # let user check segmentation result
    input_user_check = True
    while input_user_check:
        # # draw original glottal reference point
        # frame = cv2.circle(frame, (x, y), 2, [0, 0, 255], -1)
        # # draw rotated glottal reference point
        # if rotationCorrection:
        #     frame = cv2.circle(frame, (ref_rot[0], ref_rot[1]), 2, [0, 255, 0], -1)
        frame_contour = display.drawGlottisContour(frame, glottis_contour_region_growing, [0, 0, 255])
        input_user = input("Is the glottis segmentation correct? (y/n)\n")
        if input_user == "n":
            frame_result = frame.copy()
            mask_user = user.getMaskForUser(frame.shape[0], frame.shape[1])
            mask_seeds = np.zeros((frame.shape[0], frame.shape[1])).astype('uint8')
            frame[mask_user == 255] = [0, 0, 255]
            cv2.namedWindow('Input Mask', 1)
            cv2.setMouseCallback("Input Mask", callbackMouseClick)
            callback = True
            while callback:
                cv2.imshow('Input Mask', frame)
                k = cv2.waitKey(1) & 0xFF
                if k == ord('y'):
                    seed_points = user.getSeedsFromMask(mask_seeds)
                    # homogeneity_criterion = segmentation.getHomogeneityCriterion(frame_gray, seed_points)
                    # fix homogeneity criterion to value
                    homogeneity_criterion = 1.02
                    region_growing = segmentation.regionGrowing(frame_gray, seed_points, homogeneity_criterion)
                    glottis_contour = segmentation.getGlottisContourRegionGrowing(region_growing)
                    frame_contour = display.drawGlottisContour(frame_result, glottis_contour, [0, 255, 0])
                    callback = False
            input_user_check = False
            cv2.imwrite(saving_path + patient + "_" + sequence_number + "_Glottal_Segmentation_First_Frame_Reverse.png",
                        frame_contour)
            # instantiate new VideoWriter object
            output2 = cv2.VideoWriter(
                saving_path + patient + "_" + sequence_number + '_segmentation_result_reverse.mp4',
                cv2.VideoWriter_fourcc(*"mp4v"), 15.0, (256, 256))
        elif input_user == "y":
            input_user_check = False
            glottis_contour = glottis_contour_region_growing
            frame_contour = cv2.drawContours(frame, [glottis_contour_region_growing], 0, [0, 255, 0], 1)
            file.write("Homogeneity criterion: ")
            file.write(str(homogeneity_criterion))
            file.write("\n\n")
            cv2.imwrite(saving_path + patient + "_" + sequence_number + "_Glottal_Segmentation_First_Frame_Reverse.png",
                        frame_contour)
            # instantiate new VideoWriter object
            output2 = cv2.VideoWriter(
                saving_path + patient + "_" + sequence_number + '_segmentation_result_reverse.mp4',
                cv2.VideoWriter_fourcc(*"mp4v"), 15.0, (256, 256))
        else:
            print("Input faulty!")

    # calculation of (glottal area/total frame surface area) ratio in percent
    glottal_area = (cv2.contourArea(glottis_contour)/(frame.shape[0] * frame.shape[1])) * 100
    # reset lists for evolution of glottal area
    frame_number_list_area = list()
    area_list = list()

    frame_number_list_area.append(frame_number)
    area_list.append(glottal_area)

    # calculate glottal area for first frame
    area = cv2.contourArea(glottis_contour)

    # reset counter for number of unsuccessful glottis segmentation attempts
    check_mean = 0

    # processing of further frames of sequence (reverse order)
    for frame in frame_list:
        # start with frame_number = 2
        frame_number = frame_number + 1
        print(frame_number)

        if rotationCorrection:
            # rotate frame to correct glottal orientation
            frame = segmentation.rotate_frame(frame, -angle_glottal_midline)

        # preprocessing
        # bilateral filter (edge-preserving)
        frame_bilateral = prepro.bilateralFiltering(frame, 5, 75, 75)
        # transformation of RGB frame into grayscale frame
        frame_gray = prepro.convertToGray(frame_bilateral)
        # contrast enhancement
        frame_gray = prepro.enhancementContrast(frame_gray, 1.1)
        # bilateral filter (edge-preserving)
        frame_gray = prepro.bilateralFiltering(frame_gray, 3, 25, 5)

        # apply region growing

        # analyze every 4th frame only -> virtual frame rate of 1000 Hz
        if divmod(frame_number-1, 4)[1] == 0:
            # if glottis contour exists
            if not (len(glottis_contour) == 0):
                # get extremal points on glottis contour
                extLeft, extRight, extTop, extBot = segmentation.getExtremePointsContour(glottis_contour)
                # get point grid for region growing procedure (rectangular)
                mask_grid_region_growing = segmentation.getGridForRegionGrowing(frame.shape[0], frame.shape[1],
                                                                                extLeft, extRight, extTop, extBot)
                # get seed point grid for region growing procedure (adapted to glottis contour)
                seed_points = segmentation.getSeedPointsRegionGrowingFirstFrame(frame.shape[0], frame.shape[1],
                                                                                glottis_contour,
                                                                                mask_grid_region_growing)
            if check_mean < 2:
                # apply region growing method
                region_growing = segmentation.regionGrowing(frame_gray, seed_points, homogeneity_criterion)
                # region_growing = segmentation.regionGrowingRefined(frame_gray, seed_points, homogeneity_criterion)
            else:
                # abort
                region_growing = np.zeros((frame.shape[0], frame.shape[1])).astype('uint8')

            # update glottis contour based on region growing result
            glottis_contour = segmentation.getGlottisContourRegionGrowing(region_growing)

            # if no contour found
            if glottis_contour == []:
                # increment "check_mean"
                check_mean = check_mean + 1
            else:
                # if glottis area has increased by more than 50% (was 1.20) with respect to previous frame
                # (occurs when glottis closure completed)
                if cv2.contourArea(glottis_contour)/area > 1.50:
                    check_mean = check_mean + 1
                    # reset glottis contour
                    glottis_contour = []
                    # if exactly one unsuccessful segmentation attempt completed
                    if check_mean == 1:
                        # set glottis segmentation result of current frame n to result of frame (n-2)
                        glottis_contour = glottis_contour_before_before

            if not (len(glottis_contour) == 0):
                # calculation of glottal area
                try:
                    glottal_area = (cv2.contourArea(glottis_contour) / (frame.shape[0] * frame.shape[1])) * 100
                    frame_number_list_area.append(frame_number)
                    area_list.append(glottal_area)
                except:
                    pass

                # update glottis centroid coordinates
                M = cv2.moments(glottis_contour)
                # horizontal coordinate of glottis centroid
                cx = int(M['m10'] / M['m00'])
                # vertical coordinate of glottis centroid
                cy = int(M['m01'] / M['m00'])

                # calculation of glottal area
                area = cv2.contourArea(glottis_contour)
                # if fourth frame in analysis (frame_number equal to 13)
                # if second frame in analysis (frame_number equal to 5)
                if frame_number == 5:
                    # store current glottis segmentation result for comparison with future segmentations
                    glottis_contour_before = glottis_contour
                # if fifth frame in analysis or later (frame_number larger or equal 17)
                # if third frame in analysis or later (frame_number larger or equal 9)
                if frame_number >= 9:
                    # store previous glottis segmentation result for comparison with future segmentations
                    glottis_contour_before_before = glottis_contour_before
                    # store current glottis segmentation result for comparison with future segmentations
                    glottis_contour_before = glottis_contour
                frame = cv2.drawContours(frame, [glottis_contour], 0, [0, 255, 0], 1)
                # write frame to full result sequence
                output_all.write(frame)
            frame_large = cv2.resize(frame, (int(2.0 * frame.shape[1]), int(2.0 * frame.shape[0])),
                                       interpolation=cv2.INTER_LINEAR)
            cv2.imshow("Current Segmentation Result", frame_large)
            cv2.waitKey(1)
            # write frame to result sequence for inverse analysis step
            output2.write(frame)
    if output2:
        output2.release()
    if output_all:
        output_all.release()

    plot.plotArea(frame_number_list_area, area_list, saving_path + patient + "_" + sequence_number +
                  "_Reverse_Area.png")

    # identify instant of glottis opening in ms
    # (time elapsed between last frame and frame with minimum glottal area / frame rate in Hz) * 1000
    glottal_open = (last_frame_number - frame_number_list_area[np.argmin(area_list)])/4.0
    # store duration of closed glottis state in ms
    print(glottal_open - glottal_closure)
    file.write("Duration of glottal closure in ms: " + str(glottal_open - glottal_closure))
    file.close()
    # close all windows
    display.destroyWindows()
    print("Analysis completed!")
elif input_user == "n":
    file.write("\nGlottal closure achieved and glottis open at end of sequence: no")
    file.close()
    # close all windows
    display.destroyWindows()
    print("Analysis completed!")
# if glottal closure was not complete between beginning and end of sequence
else:
    print("Input faulty/glottal closure not completed in sequence! Program execution terminated.")
    file.close()
    # close all windows
    display.destroyWindows()
